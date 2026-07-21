"""Files backend local — filesystem locale (default client="local").

Builtin backend per il `client="local"` dei verbi files/dirs. Riusa
primitive `pathlib`+`os`+`shutil`+`fnmatch`+`mimetypes` di stdlib.
Nessuna dipendenza esterna.

Verbi esposti:
- `read(args)`: legge contenuto di UN file (testo o binary base64).
- `write(args)`: scrive contenuto in UN file (overwrite/append/fail).
- `find(args)`: walk + glob pattern match.
- `move(args)`: sposta/rinomina entries (vettoriale, dst_template).
- `find_dirs(args)`: walk dirs con metadata aggregati.
- `create_dirs(args)`: crea directory (vettoriale).
- `delete_dirs(args)`: rimuove directory (vettoriale, if_empty_only).

Contratto common: tutti ritornano dict con `ok: bool` + campi
verbo-specifici. Errori per-item in `failed[]`, mai silenzio
(the design guide §2.8).

Logica portata 1:1 dagli executor `read_files.py`/`write_files.py`/
`find_files.py`/`move_files.py`/`find_dirs.py`/`create_dirs.py`/
`delete_dirs.py` esistenti (13/5/2026, Q1 canonical+args).
"""
from __future__ import annotations

import base64
import datetime
import fnmatch
import hashlib
import html
import io
import json
import mimetypes
import os
import re
import shutil
import stat
import sys
import tempfile
import zipfile
import zlib
import xml.etree.ElementTree as ET
from pathlib import Path

# Lazy: il modulo move() usa platform_policy per system-file safety net.
_RUNTIME = os.environ.get("METNOS_RUNTIME") or next(
    str(p / "runtime") for p in Path(__file__).resolve().parents
    if (p / "runtime" / "config.py").is_file())
if _RUNTIME not in sys.path:
    sys.path.insert(0, _RUNTIME)

from platform_policy import is_system_file  # noqa: E402
from messages import get as _msg  # noqa: E402
from executor_helpers import vector_result  # noqa: E402
import config as _C  # noqa: E402 §7.11
# path_alias modulo riusabile (D.1, D.3). Re-export degli alias come moduli
# locali per back-compat con test esistenti che mockano backends.files.local.
from path_alias import (  # noqa: E402
    resolve_path_with_alias as _resolve_path_with_alias,
    check_mutating_path_ambiguity as _check_mutating_path_ambiguity,
    home_dir_suggestions as _home_dir_suggestions,
    normalize_input_path as _normalize_input_path,
)

# Alias bilingue IT↔EN per i path utente standard (XDG user-dirs). Quando
# l'utente IT scrive "Immagini" su un sistema con LANG=en_US la cartella
# vera e' "Pictures": senza questo mapping find_files fallisce e il planner
# Le funzioni di alias resolver vivono ora in `runtime/path_alias.py` (D.3
# refactor 22/5/2026, modulo riusabile da local.py + list_dirs.py + altri
# executor). Import sopra al modulo. Riferimenti `_*` mantengono back-compat
# con il codice del backend.


# --- read ------------------------------------------------------------------


def _decode_document_text(raw: bytes) -> str:
    """Decode testuale bounded e senza dipendenze esterne."""
    for encoding in ("utf-8-sig", "utf-16", "cp1252", "latin-1"):
        try:
            return raw.decode(encoding)
        except (UnicodeDecodeError, LookupError):
            continue
    return raw.decode("utf-8", errors="replace")


def _xml_local_name(tag: str) -> str:
    return tag.rsplit("}", 1)[-1]


def _parse_docx_text(raw: bytes) -> str:
    with zipfile.ZipFile(io.BytesIO(raw)) as archive:
        xml = archive.read("word/document.xml")
    root = ET.fromstring(xml)
    chunks: list[str] = []
    paragraph: list[str] = []
    for node in root.iter():
        name = _xml_local_name(node.tag)
        if name in {"t", "tab", "br"}:
            paragraph.append("\t" if name == "tab" else
                             "\n" if name == "br" else (node.text or ""))
        elif name == "p" and paragraph:
            text = "".join(paragraph).strip()
            if text:
                chunks.append(text)
            paragraph = []
    if paragraph:
        chunks.append("".join(paragraph).strip())
    return "\n".join(part for part in chunks if part)


def _xlsx_cell_text(cell, shared: list[str]) -> str:
    ctype = cell.attrib.get("t")
    if ctype == "inlineStr":
        return "".join((n.text or "") for n in cell.iter()
                       if _xml_local_name(n.tag) == "t")
    value = next((n.text or "" for n in cell
                  if _xml_local_name(n.tag) == "v"), "")
    if ctype == "s" and value:
        try:
            return shared[int(value)]
        except (ValueError, IndexError):
            return value
    if ctype == "b":
        return "TRUE" if value == "1" else "FALSE"
    return value


def _parse_xlsx_text(raw: bytes) -> str:
    with zipfile.ZipFile(io.BytesIO(raw)) as archive:
        names = set(archive.namelist())
        shared: list[str] = []
        if "xl/sharedStrings.xml" in names:
            root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
            for item in root.iter():
                if _xml_local_name(item.tag) == "si":
                    shared.append("".join((n.text or "") for n in item.iter()
                                          if _xml_local_name(n.tag) == "t"))
        sheets = sorted(name for name in names
                        if re.fullmatch(r"xl/worksheets/sheet\d+\.xml", name))
        rendered: list[str] = []
        for sheet_no, name in enumerate(sheets, 1):
            root = ET.fromstring(archive.read(name))
            rows: list[str] = []
            for row in root.iter():
                if _xml_local_name(row.tag) != "row":
                    continue
                cells = [_xlsx_cell_text(cell, shared) for cell in row
                         if _xml_local_name(cell.tag) == "c"]
                if cells:
                    rows.append("\t".join(cells))
            if rows:
                rendered.append(f"[Sheet {sheet_no}]\n" + "\n".join(rows))
    return "\n\n".join(rendered)


_PDF_LITERAL_RE = re.compile(rb"\((?:\\.|[^\\)])*\)")


def _decode_pdf_literal(token: bytes) -> str:
    value = token[1:-1]
    value = re.sub(rb"\\([0-7]{1,3})",
                   lambda m: bytes([int(m.group(1), 8) & 0xFF]), value)
    replacements = {
        rb"\\n": b"\n", rb"\\r": b"\r", rb"\\t": b"\t",
        rb"\\b": b"\b", rb"\\f": b"\f", rb"\\(": b"(",
        rb"\\)": b")", rb"\\\\": b"\\",
    }
    for old, new in replacements.items():
        value = value.replace(old, new)
    return _decode_document_text(value).strip()


def _parse_pdf_text(raw: bytes) -> str:
    if not raw.startswith(b"%PDF-") or b"%%EOF" not in raw[-2048:]:
        raise ValueError("invalid_or_truncated_pdf")
    sources = [raw]
    for match in re.finditer(rb"stream\r?\n(.*?)\r?\nendstream", raw, re.S):
        stream = match.group(1)
        prefix = raw[max(0, match.start() - 256):match.start()]
        if b"FlateDecode" in prefix:
            try:
                stream = zlib.decompress(stream)
            except zlib.error:
                continue
        sources.append(stream)
    chunks: list[str] = []
    for source in sources:
        for token in _PDF_LITERAL_RE.findall(source):
            text = _decode_pdf_literal(token)
            if text and any(ch.isalnum() for ch in text):
                chunks.append(text)
    # Deduplica i token ripetuti perche' il raw include anche gli stream gia'
    # aggiunti sopra; preserva l'ordine di apparizione.
    return "\n".join(dict.fromkeys(chunks))


def _auto_parse_file(path: str, source: dict | None = None) -> dict:
    """Legge e normalizza un documento locale con sola stdlib.

    Un contenitore corrotto e' un'osservazione valida (`readable=false`), non
    un fallimento I/O: la pipeline puo' segnalarlo e proseguire sugli altri.
    """
    # Keep relative reads aligned with the user workspace convention.
    abs_path = str(_normalize_input_path(path))
    try:
        raw = Path(abs_path).read_bytes()
    except FileNotFoundError:
        return {"ok": False, "path": abs_path,
                "error_code": "ERR_PATH_NOT_FOUND",
                "error": _msg("ERR_PATH_NOT_FOUND", path=abs_path)}
    except PermissionError:
        return {"ok": False, "path": abs_path,
                "error_code": "ERR_PERMISSION_DENIED",
                "error": _msg("ERR_PERMISSION_DENIED")}
    except OSError as exc:
        return {"ok": False, "path": abs_path, "error_code": "ERR_OP_FAILED",
                "error": _msg("ERR_OP_FAILED", reason=str(exc))}

    suffix = Path(abs_path).suffix.casefold()
    raw_sha = hashlib.sha256(raw).hexdigest()
    readable = True
    diagnostic = ""
    try:
        if suffix == ".docx":
            content = _parse_docx_text(raw)
            file_type = "docx"
        elif suffix == ".xlsx":
            content = _parse_xlsx_text(raw)
            file_type = "xlsx"
        elif suffix == ".pdf":
            content = _parse_pdf_text(raw)
            file_type = "pdf"
            if not content.strip():
                readable = False
                diagnostic = "pdf_text_unavailable"
        elif suffix == ".xls":
            content = ""
            file_type = "xls"
            readable = False
            diagnostic = "legacy_xls_not_supported_without_converter"
        else:
            content = _decode_document_text(raw)
            file_type = suffix.lstrip(".") or "text"
    except (ValueError, KeyError, zipfile.BadZipFile, ET.ParseError,
            UnicodeError, OSError) as exc:
        content = ""
        file_type = suffix.lstrip(".") or "unknown"
        readable = False
        diagnostic = str(exc) or type(exc).__name__

    normalized = re.sub(r"\s+", " ", content).strip().casefold()
    entry = dict(source or {})
    entry.update({
        "path": abs_path,
        "name": entry.get("name") or Path(abs_path).name,
        "ok": True,
        "content": content,
        "file_type": file_type,
        "readable": readable,
        "signature": raw_sha,
        "sha256": raw_sha,
        "content_sha256": (hashlib.sha256(normalized.encode("utf-8")).hexdigest()
                           if normalized else ""),
        "file_size": len(raw),
    })
    if diagnostic:
        entry["parse_diagnostic"] = diagnostic
    return entry


def read(args: dict) -> dict:
    """Legge il contenuto di UN file dal filesystem locale.

    Args: path (str), encoding (utf-8|latin-1|binary, default utf-8),
          max_bytes (int|None), tail_bytes (int|None), offset (int, default 0).

    `max_bytes` e `tail_bytes` sono mutuamente esclusivi. Encoding 'binary'
    ritorna content come base64. Truncation visibility (§2.7+§2.11): se la
    lettura non copre l'intero file, espone truncated/used/available_total/
    cap_field/cap_value.

    Vettoriale §2.1: con `paths=[...]` o `entries` (da from_step, proietta
    entries[*].path) legge N file e ritorna `entries=[{path, content, ...}]`
    (§2.6: read arricchisce una lista). La forma scalare `path` resta {ok,
    content, metadata} (back-compat).
    """
    # Espansione DIRECTORY → file dentro (universale: "leggi i file in una
    # cartella" in 1 call deterministica, niente find_files+read_files a 2
    # passi). `pattern` (default '*') filtra; ricorsione no. Vale per path
    # scalare-dir e per ogni elemento-dir di paths.
    import glob as _glob

    def _expand_dir(p):
        if not isinstance(p, str) or not p:
            return []
        ap = os.path.abspath(os.path.expanduser(p))
        if os.path.isdir(ap):
            pat = args.get("pattern") or "*"
            return sorted(f for f in _glob.glob(os.path.join(ap, pat))
                          if os.path.isfile(f))
        return [p]

    _vec = None
    _source_by_path: dict[str, dict] = {}
    _ps = args.get("paths")
    _es = args.get("entries")
    _scalar = args.get("path")
    if isinstance(_ps, list):
        _vec = []
        for p in _ps:
            _vec.extend(_expand_dir(p))
    elif isinstance(_es, list):
        _vec = []
        for e in _es:
            if isinstance(e, dict) and isinstance(e.get("path"), str) and e.get("path"):
                _vec.extend(_expand_dir(e["path"]))
                _source_by_path[os.path.abspath(os.path.expanduser(e["path"]))] = e
    elif isinstance(_scalar, str) and os.path.isdir(
            os.path.abspath(os.path.expanduser(_scalar))):
        _vec = _expand_dir(_scalar)  # path scalare = directory → vettoriale
    if _vec is not None:
        try:
            _maxf = int(args.get("max_files") or _DEFAULT_MAX_FILES)
        except (TypeError, ValueError):
            _maxf = _DEFAULT_MAX_FILES
        _avail = len(_vec)
        _trunc = _maxf > 0 and len(_vec) > _maxf
        if _trunc:
            _vec = _vec[:_maxf]
        _base = {k: v for k, v in args.items()
                 if k not in ("paths", "entries", "parse")}
        _parse = args.get("parse")
        out_entries = []
        ok_count = fail_count = 0
        for _p in _vec:
            if _parse == "auto":
                r = _auto_parse_file(
                    _p, _source_by_path.get(os.path.abspath(os.path.expanduser(_p))))
            else:
                r = read({**_base, "path": _p})  # delega alla logica scalare
            if r.get("ok"):
                ent = {"path": _p, "ok": True}
                # parse="json": fonde i campi del JSON nell'entry (record
                # interrogabile da filter_entries) — simmetrico al write JSON
                # dell'inbound. Default: content come stringa + metadata.
                if _parse == "auto":
                    ent = r
                elif _parse == "json":
                    try:
                        _pj = json.loads(r.get("content") or "")
                        if isinstance(_pj, dict):
                            ent.update(_pj)
                        else:
                            ent["content"] = _pj
                    except (ValueError, TypeError):
                        ent["content"] = r.get("content")
                else:
                    ent["content"] = r.get("content")
                    ent.update(r.get("metadata") or {})
                ok_count += 1
            else:
                ent = {"path": _p, "ok": False,
                       "error_code": r.get("error_code"), "error": r.get("error")}
                fail_count += 1
            out_entries.append(ent)
        duplicates: list[dict] = []
        if args.get("deduplicate_content") and out_entries:
            unique: list[dict] = []
            by_signature: dict[str, dict] = {}
            by_content: dict[str, dict] = {}
            for entry in out_entries:
                if not entry.get("ok"):
                    unique.append(entry)
                    continue
                raw_key = str(entry.get("sha256") or "")
                content_key = str(entry.get("content_sha256") or "")
                original = (by_signature.get(raw_key) if raw_key else None)
                if original is None and content_key:
                    original = by_content.get(content_key)
                if original is None:
                    unique.append(entry)
                    if raw_key:
                        by_signature[raw_key] = entry
                    if content_key:
                        by_content[content_key] = entry
                    continue
                duplicate = {
                    "path": entry.get("path"),
                    "duplicate_of": original.get("path"),
                    "signature": raw_key,
                    "content_sha256": content_key,
                }
                duplicates.append(duplicate)
                original.setdefault("duplicate_paths", []).append(entry.get("path"))
                original["duplicate_count"] = len(original["duplicate_paths"])
            out_entries = unique
        input_ok_count = ok_count
        if _parse == "auto" and args.get("deduplicate_content"):
            ok_count = sum(1 for entry in out_entries if entry.get("ok"))
        out = {"ok": fail_count == 0, "ok_count": ok_count,
               "fail_count": fail_count, "entries": out_entries}
        if _parse == "auto":
            out.update({
                "duplicates": duplicates,
                "deduped_count": len(duplicates),
                "input_ok_count": input_ok_count,
                "readable_count": sum(1 for e in out_entries
                                      if e.get("ok") and e.get("readable")),
                "unreadable_count": sum(1 for e in out_entries
                                        if e.get("ok") and not e.get("readable")),
            })
        if fail_count:
            _ff = next((e for e in out_entries if not e.get("ok")), None)
            if _ff:
                out["error_code"] = _ff.get("error_code")
                out["error"] = _ff.get("error")
        if _trunc:
            out.update({"truncated": True, "truncated_what": "files",
                        "used": len(_vec), "available_total": _avail,
                        "cap_field": "max_files", "cap_value": _maxf})
        return out

    path = args.get("path")
    encoding = args.get("encoding", "utf-8")
    max_bytes = args.get("max_bytes")
    tail_bytes = args.get("tail_bytes")
    offset = args.get("offset", 0)

    # Robustezza §2.4: 0 come placeholder = None (no limit).
    if max_bytes == 0:
        max_bytes = None
    if tail_bytes == 0:
        tail_bytes = None

    if not path:
        return {"ok": False, "error_code": "ERR_ARG_MISSING",
                "error": _msg("ERR_ARG_MISSING", arg="path")}
    if max_bytes is not None and tail_bytes is not None:
        return {"ok": False, "error_code": "ERR_ARG_INVALID",
                "error": _msg("ERR_ARG_INVALID", arg="max_bytes/tail_bytes", reason="mutuamente esclusivi")}

    # Relative reads use the same user-workspace convention as discovery.
    abs_path = str(_normalize_input_path(path))

    try:
        file_size = os.path.getsize(abs_path)

        if tail_bytes is not None:
            seek_to = max(0, file_size - tail_bytes)
            read_n = tail_bytes
            mode_str = "tail"
        else:
            seek_to = offset
            read_n = max_bytes
            mode_str = "offset" if offset > 0 else ("head" if max_bytes else "full")

        will_truncate = (read_n is not None) and (seek_to + (read_n or 0) < file_size)

        if encoding == "binary":
            with open(abs_path, "rb") as f:
                if seek_to:
                    f.seek(seek_to)
                data = f.read(read_n) if read_n else f.read()
            out = {
                "ok": True,
                "content": base64.b64encode(data).decode("ascii"),
                "metadata": {
                    "encoding": "binary-base64",
                    "bytes": len(data),
                    "path": abs_path,
                    "file_size": file_size,
                    "read_offset": seek_to,
                    "read_mode": mode_str,
                },
            }
            if will_truncate:
                out["truncated"] = True
                out["truncated_what"] = "byte"
                out["used"] = len(data)
                out["available_total"] = file_size
                out["cap_field"] = "max_bytes"
                out["cap_value"] = read_n
            return out
        else:
            with open(abs_path, "rb") as f:
                if seek_to:
                    f.seek(seek_to)
                raw = f.read(read_n) if read_n else f.read()
            try:
                text = raw.decode(encoding)
            except UnicodeDecodeError:
                text = raw.decode(encoding, errors="replace")
            out = {
                "ok": True,
                "content": text,
                "metadata": {
                    "encoding": encoding,
                    "bytes": len(raw),
                    "chars": len(text),
                    "path": abs_path,
                    "file_size": file_size,
                    "read_offset": seek_to,
                    "read_mode": mode_str,
                },
            }
            if will_truncate:
                out["truncated"] = True
                out["truncated_what"] = "byte"
                out["used"] = len(raw)
                out["available_total"] = file_size
                out["cap_field"] = "max_bytes"
                out["cap_value"] = read_n
            return out
    except FileNotFoundError:
        return {"ok": False, "error_code": "ERR_PATH_NOT_FOUND",
                "error": _msg("ERR_PATH_NOT_FOUND", path=str(abs_path))}
    except PermissionError:
        return {"ok": False, "error_code": "ERR_PERMISSION_DENIED",
                "error": _msg("ERR_PERMISSION_DENIED"), "detail": f"path outside allowed scope: {abs_path}"}
    except IsADirectoryError:
        return {"ok": False, "error_code": "ERR_PATH_WRONG_TYPE",
                "error": _msg("ERR_PATH_WRONG_TYPE", expected="file", actual="directory", path=str(abs_path))}
    except OSError as e:
        return {"ok": False, "error_code": "ERR_OP_FAILED",
                "error": _msg("ERR_OP_FAILED", reason=f"os error: {e}")}


# --- write -----------------------------------------------------------------


_VALID_WRITE_MODES = ("overwrite", "append", "fail_if_exists", "skip_if_exists")
_DEFAULT_MAX_FILES = 500


def _safe_format(template: str, entry: dict):
    """`template.format(**entry)` deterministico §7.9. Ritorna (ok, valore_o_errore).
    Campo mancante o spec invalido → (False, messaggio onesto), niente raise."""
    # §2.4: l'LLM usa spesso la sintassi ${entry.campo} o ${campo} (convenzione
    # di piping) invece del {campo} di str.format → normalizza prima (bug q28
    # 5/6). Lascia intatti i ${X:Y} con due punti (es. ${RUNTIME:..} già risolti).
    if isinstance(template, str) and "${" in template:
        import re as _re
        template = _re.sub(r"\$\{\s*entry\.(\w+)\s*\}", r"{\1}", template)
        template = _re.sub(r"\$\{\s*(\w+)\s*\}", r"{\1}", template)
    try:
        # §2.8: un campo del template ASSENTE nell'entry → stringa vuota, NON
        # hard-fail. L'LLM sceglie i nomi-campo best-effort (es. {memory} ma
        # l'entry processo ha 'mem_pct') → rendi i campi presenti, lascia vuoti
        # i mancanti (bug q28/q34 5/6: il write falliva tutto per un nome
        # leggermente diverso). Spec malformato → errore onesto.
        class _SafeDict(dict):
            def __missing__(self, _k):
                return ""
        flat = entry if isinstance(entry, dict) else {"value": entry}
        return True, template.format_map(_SafeDict(flat))
    except (ValueError, TypeError, IndexError, AttributeError) as ex:
        return False, _msg("ERR_ARG_INVALID", arg="template",
                            reason=f"template non valido: {ex}")


def _derive_content(entry, content_field, content_template, content_format):
    """Deriva il contenuto (stringa) di un file da una `entry`. Precedenza:
    content_field > content_template > content_format(text) > JSON (default).
    Deterministico §7.9. Ritorna (ok, valore_o_errore)."""
    if content_field:
        if isinstance(entry, dict) and content_field in entry:
            v = entry[content_field]
            return True, (v if isinstance(v, str)
                          else json.dumps(v, ensure_ascii=False, indent=2))
        return False, _msg("ERR_ARG_INVALID", arg="content_field",
                            reason=f"'{content_field}' assente nell'entry")
    if content_template:
        return _safe_format(
            content_template,
            entry if isinstance(entry, dict) else {"value": entry})
    if content_format == "text":
        return True, (entry if isinstance(entry, str)
                      else json.dumps(entry, ensure_ascii=False, indent=2))
    return True, json.dumps(entry, ensure_ascii=False, indent=2)


def _collect_write_specs(args: dict):
    """Normalizza gli input in una lista di spec {path, content, encoding,
    mode}. Vettoriale §2.1: la lista ammette anche un solo elemento. Shape
    in ordine di precedenza: files[] esplicito → entries[]+path_template →
    paths[]+contents[] → path+content scalare. Ritorna (specs, errore_o_None)."""
    enc_default = args.get("encoding", "utf-8")
    mode_default = args.get("mode", "overwrite")
    specs: list[dict] = []

    files = args.get("files")
    entries = args.get("entries")
    paths = args.get("paths")
    contents = args.get("contents")

    if isinstance(files, list):
        for i, f in enumerate(files):
            if not isinstance(f, dict) or not f.get("path"):
                return None, _msg("ERR_ARG_INVALID", arg="files",
                                  reason=f"elemento {i} senza 'path'")
            if f.get("content") is None:
                return None, _msg("ERR_ARG_INVALID", arg="files",
                                  reason=f"elemento {i} senza 'content'")
            specs.append({"path": f["path"], "content": f["content"],
                          "encoding": f.get("encoding", enc_default),
                          "mode": f.get("mode", mode_default)})
        return specs, None

    if isinstance(entries, list):
        path_template = args.get("path_template")
        # §2.8: un `path_template` SENZA placeholder (`{campo}`/`${campo}`) NON è
        # un template per-entry — è un literal. Il planner ci mette per errore il
        # CONTENUTO (es. path_template="RISC-V è ...") e, avendo precedenza sul
        # `path` scalare, scriverebbe un file mal-nominato nel cwd invece che nel
        # path dato. Scartalo → si ricade sull'AGGREGATO verso `path` (bug
        # latente write path=contenuto, q4/q27). Deterministico §7.9.
        if isinstance(path_template, str) and "{" not in path_template:
            path_template = None
        content_field = args.get("content_field")
        content_template = args.get("content_template")
        content_format = args.get("content_format")
        # set_fields: campi costanti fusi in OGNI entry prima di serializzare
        # (read-modify-write: es. set_fields={"status":"answered"} per marcare
        # le issue gestite). Universale: persisti la lista con un campo aggiornato.
        set_fields = args.get("set_fields")
        # AGGREGATO (§2.10, 4/6/2026): entries + `path` SCALARE senza
        # `path_template` → UN solo file con TUTTE le entries serializzate
        # (es. extract_entries → write_files(path="/dir/out.txt", from_step=N)).
        # Senza questo, il caso "salva la lista in un unico file" falliva con
        # "path_template/content mancante" (fix q4 live-test).
        scalar_path = args.get("path")
        if not path_template and scalar_path and str(scalar_path).strip():
            parts = []
            for entry in entries:
                if isinstance(set_fields, dict) and isinstance(entry, dict):
                    entry = {**entry, **set_fields}
                ok, c = _derive_content(entry, content_field,
                                        content_template, content_format)
                if not ok:
                    return None, c
                parts.append(c if isinstance(c, str) else str(c))
            specs.append({"path": scalar_path, "content": "\n".join(parts),
                          "encoding": enc_default, "mode": mode_default})
            return specs, None
        _explicit_content = bool(content_field or content_template
                                 or content_format)
        for entry in entries:
            if isinstance(set_fields, dict) and isinstance(entry, dict):
                entry = {**entry, **set_fields}
            if path_template:
                ok, p = _safe_format(
                    path_template,
                    entry if isinstance(entry, dict) else {"value": entry})
                if not ok:
                    return None, p
                ok, c = _derive_content(entry, content_field,
                                        content_template, content_format)
                if not ok:
                    return None, c
            elif isinstance(entry, dict) and entry.get("path"):
                # §2.8/§2.9: `entry["path"]` localizza un file GIA' ESISTENTE
                # (output di un producer: i `results` di create_files_spreadsheet
                # portano path=<xlsx creato>). Riusarlo come OUTPUT scrivendoci il
                # DEFAULT json.dumps(entry) CORROMPE il file (bug xlsx-clobber,
                # turn 3da933e5). Consentito SOLO con una fonte di contenuto
                # esplicita, o se l'entry porta un campo `content` genuino
                # (write-spec {path,content}); altrimenti è un mis-pipe (§2.10)
                # → errore onesto, MAI clobber. §7.9 deterministico.
                if _explicit_content:
                    ok, c = _derive_content(entry, content_field,
                                            content_template, content_format)
                    if not ok:
                        return None, c
                elif "content" in entry:
                    v = entry["content"]
                    c = (v if isinstance(v, str)
                         else json.dumps(v, ensure_ascii=False, indent=2))
                else:
                    return None, _msg(
                        "ERR_ARG_INVALID", arg="entries",
                        reason="entries con 'path' verso file esistenti ma senza "
                               "contenuto: per creare NUOVI file usa "
                               "path_template, per scrivere un campo usa "
                               "content_field")
                p = entry["path"]
            else:
                return None, _msg("ERR_ARG_MISSING", arg="path_template")
            specs.append({"path": p, "content": c,
                          "encoding": enc_default, "mode": mode_default})
        return specs, None

    if isinstance(paths, list) and isinstance(contents, list):
        if len(paths) != len(contents):
            return None, _msg("ERR_ARG_INVALID", arg="contents",
                              reason="paths e contents hanno lunghezza diversa")
        for p, c in zip(paths, contents):
            specs.append({"path": p, "content": c,
                          "encoding": enc_default, "mode": mode_default})
        return specs, None

    # Scalare: degenere N=1 (la "lista" ha un solo elemento).
    path = args.get("path")
    content = args.get("content")
    if path is None or not str(path).strip():
        return None, _msg("ERR_ARG_MISSING", arg="path")
    if content is None:
        return None, _msg("ERR_ARG_MISSING", arg="content")
    # content NON-stringa (§2.10, 4/6/2026): il planner spesso pipa una LISTA o
    # un record come `content` (es. content={{stepN.entries}} risolto a lista) →
    # serializza in testo leggibile invece di passare un oggetto a file.write()
    # (TypeError: write() argument must be str, not list). Fix q4 live-test.
    if not isinstance(content, (str, bytes)):
        content = _serialize_content(content)
    specs.append({"path": path, "content": content,
                  "encoding": enc_default, "mode": mode_default})
    return specs, None


def _serialize_content(c):
    """Serializza un content non-stringa (lista/record/scalare) in testo
    leggibile (§2.10). Lista di record monocampo → valori; record multi-campo
    o annidati → JSON; scalari → str."""
    import json as _json
    if isinstance(c, list):
        parts = []
        for x in c:
            if isinstance(x, str):
                parts.append(x)
            elif isinstance(x, dict):
                if len(x) == 1:
                    parts.append(str(next(iter(x.values()))))
                else:
                    parts.append(_json.dumps(x, ensure_ascii=False))
            else:
                parts.append(str(x))
        return "\n".join(parts)
    if isinstance(c, dict):
        return _json.dumps(c, ensure_ascii=False)
    return str(c)


def _write_one(path, content, encoding, mode):
    """Scrive UN file. Ritorna (ok, entry_o_errordict, created_parents).
    Estratto dal write() scalare per riuso nella forma vettoriale."""
    ambig = _check_mutating_path_ambiguity(path, target_must_exist=False)
    if ambig is not None:
        return False, dict(ambig, path=path), []

    # Mutating relative paths must share the reader workspace convention.
    # Otherwise a remote executor writes into its disposable sandbox CWD.
    abs_path = str(_normalize_input_path(path))
    pre_existed = os.path.exists(abs_path)

    if mode == "fail_if_exists" and pre_existed:
        return False, {"path": abs_path, "ok": False,
                       "error_code": "ERR_DST_EXISTS",
                       "error": _msg("ERR_DST_EXISTS"),
                       "detail": str(abs_path)}, []
    if mode == "skip_if_exists" and pre_existed:
        # Idempotenza deterministica (dedup): file gia' presente → no-op
        # onesto (created=False, skipped=True), non un errore.
        return True, {"path": abs_path, "created": False, "skipped": True,
                      "bytes_written": 0, "encoding": encoding,
                      "mode": mode}, []

    # mkdir -p del parent + registro per undo (§2.4 robustezza NL→det).
    parent = os.path.dirname(abs_path)
    created_parents: list[str] = []
    if parent and not os.path.isdir(parent):
        chain: list[str] = []
        p = parent
        while p and not os.path.isdir(p):
            chain.append(p)
            np = os.path.dirname(p)
            if np == p:
                break
            p = np
        try:
            os.makedirs(parent, exist_ok=True)
            created_parents = list(reversed(chain))
        except OSError as e:
            return False, {"path": abs_path, "ok": False,
                           "error_code": "ERR_PARENT_MKDIR_FAIL",
                           "error": _msg("ERR_PARENT_MKDIR_FAIL",
                                         path=str(parent), reason=str(e))}, []

    try:
        if encoding == "binary":
            data = base64.b64decode(content)
            with open(abs_path, "ab" if mode == "append" else "wb") as f:
                f.write(data)
            entry = {"path": abs_path, "created": (not pre_existed),
                     "bytes_written": len(data), "encoding": "binary",
                     "mode": mode}
            return True, entry, created_parents
        else:
            with open(abs_path, "a" if mode == "append" else "w",
                      encoding=encoding) as f:
                f.write(content)
            entry = {"path": abs_path, "created": (not pre_existed),
                     "bytes_written": len(content.encode(encoding)),
                     "encoding": encoding, "mode": mode}
            return True, entry, created_parents
    except PermissionError:
        return False, {"path": abs_path, "ok": False,
                       "error_code": "ERR_PERMISSION_DENIED",
                       "error": _msg("ERR_PERMISSION_DENIED"),
                       "detail": f"path outside allowed scope: {abs_path}"}, []
    except IsADirectoryError:
        return False, {"path": abs_path, "ok": False,
                       "error_code": "ERR_PATH_WRONG_TYPE",
                       "error": _msg("ERR_PATH_WRONG_TYPE", expected="file",
                                     actual="directory", path=str(abs_path))}, []
    except OSError as e:
        return False, {"path": abs_path, "ok": False,
                       "error_code": "ERR_OP_FAILED",
                       "error": _msg("ERR_OP_FAILED", reason=f"os error: {e}")}, []
    except Exception as e:
        return False, {"path": abs_path, "ok": False,
                       "error_code": "ERR_OP_FAILED",
                       "error": _msg("ERR_OP_FAILED",
                                     reason=f"unexpected {type(e).__name__}: {e}")}, []


def write(args: dict) -> dict:
    """Scrive contenuto in uno o piu' file. Vettoriale §2.1: la lista ammette
    anche un solo elemento (degenere N=1). Shape input (precedenza):
      - files=[{path, content, encoding?, mode?}, ...]   vettore esplicito;
      - entries=[...] + path_template (+ content_field | content_template |
        content_format)   persiste record piped via from_step come file;
      - paths=[...] + contents=[...]   array paralleli;
      - path + content   scalare (N=1).
    Output (§2.6 trasformativo → results): {ok, ok_count, fail_count,
    created_count, skipped_count, results:[...], dirs_created:[...]}.
    `mode='skip_if_exists'` rende la scrittura idempotente (dedup)."""
    mode_default = args.get("mode", "overwrite")
    if mode_default not in _VALID_WRITE_MODES:
        return {"ok": False, "error_code": "ERR_ARG_INVALID",
                "error": _msg("ERR_ARG_INVALID", arg="mode",
                              reason=f"invalid value '{mode_default}'")}

    specs, err = _collect_write_specs(args)
    if err is not None:
        return {"ok": False, "error_code": "ERR_ARG_INVALID", "error": err}
    for sp in specs:
        if sp["mode"] not in _VALID_WRITE_MODES:
            return {"ok": False, "error_code": "ERR_ARG_INVALID",
                    "error": _msg("ERR_ARG_INVALID", arg="mode",
                                  reason=f"invalid value '{sp['mode']}'")}

    # Cap superiore esplicito §2.1/§2.7.
    try:
        max_files = int(args.get("max_files") or _DEFAULT_MAX_FILES)
    except (TypeError, ValueError):
        max_files = _DEFAULT_MAX_FILES
    available_total = len(specs)
    truncated = max_files > 0 and len(specs) > max_files
    if truncated:
        specs = specs[:max_files]

    results: list[dict] = []
    dirs_created: list[str] = []
    ok_count = fail_count = created_count = skipped_count = 0
    for sp in specs:
        ok, entry, parents = _write_one(
            sp["path"], sp["content"], sp["encoding"], sp["mode"])
        results.append(entry)
        if ok:
            ok_count += 1
            if entry.get("skipped"):
                skipped_count += 1
            elif entry.get("created"):
                created_count += 1
            for d in parents:
                if d not in dirs_created:
                    dirs_created.append(d)
        else:
            fail_count += 1

    out = {
        "ok": fail_count == 0,
        "ok_count": ok_count,
        "fail_count": fail_count,
        "created_count": created_count,
        "skipped_count": skipped_count,
        "results": results,
        "dirs_created": dirs_created,
    }
    # §2.8: superficie l'errore al top-level (parità con la forma scalare
    # storica: error_code/error/detail) cosi' la sintesi error-first del
    # runtime e i matcher non perdono il fallimento dentro results[].
    if fail_count:
        first_fail = next((r for r in results if isinstance(r, dict)
                           and r.get("ok") is False), None)
        if first_fail:
            out["error_code"] = first_fail.get("error_code")
            out["error"] = first_fail.get("error")
            if first_fail.get("detail"):
                out["detail"] = first_fail["detail"]
            # §2.10: su ERR_AMBIGUOUS_PATH lifta anche `candidates` a top-level
            # (parità con gli altri executor): il runtime costruisce il dialog
            # di disambiguazione dal top-level, non da results[0].
            if first_fail.get("candidates") is not None:
                out["candidates"] = first_fail["candidates"]
    # §2.7 truncation visibility.
    if truncated:
        out.update({"truncated": True, "truncated_what": "files",
                    "used": len(specs), "available_total": available_total,
                    "cap_field": "max_files", "cap_value": max_files})
    return out


# --- find ------------------------------------------------------------------

_KIND_PREFIX = {
    "image": ("image/",),
    "video": ("video/",),
    "audio": ("audio/",),
    "text": ("text/",),
    "document": ("application/pdf", "application/msword",
                  "application/vnd.openxmlformats-officedocument",
                  "application/vnd.oasis.opendocument",
                  "application/rtf"),
    "archive": ("application/zip", "application/x-tar", "application/gzip",
                 "application/x-7z-compressed", "application/x-rar"),
}


def _mime_for(name: str) -> str:
    mt, _ = mimetypes.guess_type(name.lower())
    return mt or "application/octet-stream"


def _kind_for(mime: str) -> str:
    for kind, prefixes in _KIND_PREFIX.items():
        if any(mime.startswith(pref) for pref in prefixes):
            return kind
    return "binary"


def _parse_compound_pattern(value):
    """Accetta str o list[str] e ritorna list[str] di pattern atomici.
    Una str con separatori naturali (virgola, pipe) viene splittata.
    """
    out: list[str] = []
    if value is None:
        return out
    if isinstance(value, list):
        for v in value:
            out.extend(_parse_compound_pattern(v))
        return out
    if isinstance(value, str):
        s = value.strip()
        if not s:
            return out
        if any(sep in s for sep in (",", "|")):
            parts = re.split(r"[,|]", s)
        else:
            parts = [s]
        for p in parts:
            p = p.strip()
            if p:
                out.append(p)
    return out


def find(args: dict) -> dict:
    """Cerca file per pattern dentro base_path. Args: base_path, pattern|patterns, ..."""
    base_path = args.get("base_path")
    recursive = args.get("recursive", True)
    max_results = args.get("max_results", 1000)
    max_depth = args.get("max_depth", 10)
    include_dirs = args.get("include_dirs", False)
    case_sensitive = args.get("case_sensitive", False)

    patterns = _parse_compound_pattern(args.get("pattern")) + _parse_compound_pattern(args.get("patterns"))

    # Robustezza §2.4 (22/5/2026): se patterns non e' specificato, default
    # ["*"] (= "tutti i file"). Caso live turn 7580b454: planner chiama
    # find_files({base_path: ...}) per query "quanti file" senza patterns;
    # fallisce ERR_ARG_MISSING e ripiega su list_dirs (top-level only).
    if not patterns:
        patterns = ["*"]
    if not base_path:
        return {"ok": False, "error_code": "ERR_ARG_MISSING",
                "error": _msg("ERR_ARG_MISSING", arg="base_path")}
    if not isinstance(max_results, int) or max_results < 1:
        return {"ok": False, "error_code": "ERR_ARG_INVALID",
                "error": _msg("ERR_ARG_INVALID", arg="max_results", reason="must be a positive integer")}
    if not isinstance(max_depth, int) or max_depth < 0:
        return {"ok": False, "error_code": "ERR_ARG_INVALID",
                "error": _msg("ERR_ARG_INVALID", arg="max_depth", reason="must be >= 0")}

    base, alias_note = _resolve_path_with_alias(base_path)
    if not base.exists():
        # Suggerisci cartelle home esistenti: il planner puo' chiedere
        # all'utente quale intendeva, evitando loop_break generico.
        _sugg = _home_dir_suggestions(base.name)
        out = {"ok": False, "error_code": "ERR_PATH_NOT_FOUND",
               "error": _msg("ERR_PATH_NOT_FOUND", path=str(base)),
               "suggested_paths": _sugg}
        # §2.11 errore-runtime→form: il path non esiste MA ci sono candidati
        # plausibili → emetti il segnale `disambiguation` (l'executor conosce il
        # dominio: i path home esistenti). Il dispatch lo traduce in form
        # get_inputs; alla scelta ri-esegue la query con base_path=<scelto>.
        if _sugg:
            out["disambiguation"] = {
                "prompt": f"Il percorso «{base}» non esiste. Quale intendevi?",
                "var": "base_path",
                "options": [{"value": p, "label": p} for p in _sugg],
                "rerun": True,
            }
        return out
    if not base.is_dir():
        return {"ok": False, "error_code": "ERR_PATH_WRONG_TYPE",
                "error": _msg("ERR_PATH_WRONG_TYPE", expected="directory", actual="file", path=str(base))}

    def name_matches(name: str) -> bool:
        if case_sensitive:
            return any(fnmatch.fnmatchcase(name, p) for p in patterns)
        nlower = name.lower()
        return any(fnmatch.fnmatchcase(nlower, p.lower()) for p in patterns)

    def _walk_limited():
        """Itera senza scendere oltre ``max_depth``.

        ``Path.rglob`` non consente di potare l'albero: il vecchio controllo
        sulla profondita' scartava i risultati troppo profondi, ma visitava
        comunque tutto il sottoalbero (molto costoso su home/OneDrive/NAS).
        ``os.walk(topdown=True)`` permette invece di svuotare ``dirnames``
        prima della discesa. L'ordinamento rende stabile il batch fra OS.
        """
        if not recursive:
            yield from base.iterdir()
            return

        def _raise_walk_error(error):
            raise error

        for root, dirnames, filenames in os.walk(
                base, topdown=True, onerror=_raise_walk_error,
                followlinks=False):
            dirnames.sort()
            filenames.sort()
            root_path = Path(root)
            try:
                root_depth = len(root_path.relative_to(base).parts)
            except ValueError:
                continue
            child_depth = root_depth + 1
            if child_depth <= max_depth:
                for name in dirnames:
                    yield root_path / name
                for name in filenames:
                    yield root_path / name
            if child_depth >= max_depth:
                dirnames.clear()

    walker = _walk_limited()
    entries: list[dict] = []
    truncated = False
    visited = 0

    try:
        for p in walker:
            visited += 1
            try:
                depth = len(p.relative_to(base).parts)
            except ValueError:
                continue
            if depth > max_depth:
                continue
            # Il nome e' disponibile senza syscall. Su pattern selettivi
            # (es. PDF/DOCX/XLSX) evita is_dir/stat per la grande maggioranza
            # delle entry attraversate.
            if not name_matches(p.name):
                continue
            try:
                is_link = p.is_symlink()
                is_dir = p.is_dir() and not is_link
            except OSError:
                continue
            ftype = "symlink" if is_link else ("dir" if is_dir else "file")
            if ftype != "file" and not include_dirs:
                continue
            if ftype == "file":
                mime = _mime_for(p.name)
                kind = _kind_for(mime)
            elif ftype == "dir":
                mime = ""
                kind = "dir"
            else:
                mime = ""
                kind = "symlink"
            try:
                st = p.lstat() if is_link else p.stat()
                size = int(st.st_size)
                mtime = float(st.st_mtime)
            except OSError:
                size = 0
                mtime = 0.0
            entries.append({
                "path": str(p),
                "name": p.name,
                "type": ftype,
                "mime": mime,
                "kind": kind,
                "size": size,
                "mtime": mtime,
            })
            if len(entries) >= max_results:
                truncated = True
                break
    except PermissionError as e:
        return {"ok": False, "error_code": "ERR_PERMISSION_DENIED",
                "error": _msg("ERR_PERMISSION_DENIED"), "detail": str(e)}
    except OSError as e:
        return {"ok": False, "error_code": "ERR_OP_FAILED",
                "error": _msg("ERR_OP_FAILED", reason=f"os error: {e}")}

    # Sondaggio post-cap §2.11
    extra_matches = 0
    if truncated:
        # Probe esteso (22/5/2026): per query count-style ("quanti file in X")
        # con max_results piccolo (1000) ma corpus grande (es. NAS 33K+),
        # serve un sondaggio profondo per available_total veritiero.
        # Floor 100k per ~secondi su NAS lenti; senza pattern (`*`) il caller
        # vuole comunque un count globale, quindi paghiamo lo scan.
        probe_cap = max(100 * max_results, max_results + 100000)
        try:
            for p in walker:
                visited += 1
                try:
                    depth = len(p.relative_to(base).parts)
                except ValueError:
                    continue
                if depth > max_depth:
                    continue
                if not name_matches(p.name):
                    continue
                try:
                    is_link = p.is_symlink()
                    is_dir = p.is_dir() and not is_link
                except OSError:
                    continue
                ftype = "symlink" if is_link else ("dir" if is_dir else "file")
                if ftype != "file" and not include_dirs:
                    continue
                extra_matches += 1
                if extra_matches >= probe_cap:
                    break
        except (PermissionError, OSError):
            pass

    matches = [e["path"] for e in entries]
    out = {
        "ok": True,
        "entries": entries,
        "matches": matches,
        "metadata": {
            "base_path": str(base),
            "patterns": patterns,
            "recursive": recursive,
            "case_sensitive": case_sensitive,
            "count": len(entries),
            "visited": visited,
            "truncated": truncated,
            # Se il path originale non esisteva ma e' stato risolto via alias
            # bilingue IT/EN, lo segnaliamo nei metadata (planner + UI).
            **({"alias_resolved": alias_note} if alias_note else {}),
        },
    }
    if truncated:
        out["truncated"] = True
        out["truncated_what"] = "file"
        out["used"] = len(entries)
        out["cap_field"] = "max_results"
        out["cap_value"] = max_results
        out["available_total"] = len(entries) + extra_matches
    return out


# --- move ------------------------------------------------------------------


def _entry_fields(entry, src_path):
    name = entry.get("name") or src_path.name
    if "." in name and not name.startswith("."):
        stem, ext = name.rsplit(".", 1)
    else:
        stem, ext = name, ""
    parent = entry.get("parent") or str(src_path.parent)
    mtime_epoch = entry.get("mtime_epoch")
    if mtime_epoch is None:
        try:
            mtime_epoch = src_path.stat().st_mtime
        except OSError:
            mtime_epoch = 0
    dt = datetime.datetime.fromtimestamp(float(mtime_epoch))
    de = entry.get("date_epoch")
    if de is None:
        date_dt = dt
    else:
        try:
            date_dt = datetime.datetime.fromtimestamp(float(de))
        except (TypeError, ValueError, OSError):
            date_dt = dt
    place = entry.get("place") or "unknown"
    if not isinstance(place, str) or not place.strip():
        place = "unknown"
    # Passthrough dei campi STRINGA extra dell'entry (§2.4): il dst_template
    # può referenziarli (es. entries=[{src, dst}] + "{dst}" — batch restore
    # del reverse remoto, 6/7). I campi calcolati qui sotto NON sono
    # sovrascrivibili (vincono sempre).
    extra = {k: v for k, v in entry.items()
             if isinstance(v, str) and k not in (
                 "path", "name", "stem", "ext", "parent", "place")}
    return {
        **extra,
        "path": str(src_path),
        "name": name,
        "stem": stem,
        "ext": ext,
        "parent": parent,
        "date_year": date_dt.strftime("%Y"),
        "date_yymmdd": date_dt.strftime("%y%m%d"),
        "date_yyyymmdd": date_dt.strftime("%Y%m%d"),
        "place": place,
    }


def _move_one(src_path, dst_path, overwrite, parents, copy=False):
    """Esegue lo spostamento (o la COPIA con copy=True); ritorna
    (ok, error, dirs_created). copy=True: src resta (usato dal reverse
    remoto restore_blob_backup — un blob DEDUPLICATO serve più path,
    consumarlo al primo restore rompeva i duplicati; bug live 6/7)."""
    if not src_path.exists():
        return False, f"src not found: {src_path}", []
    if str(dst_path) == str(src_path):
        return False, "src and dst are the same path", []
    if dst_path.exists():
        if not overwrite:
            return False, f"dst already exists (use overwrite=true): {dst_path}", []
        try:
            if dst_path.is_dir() and not dst_path.is_symlink():
                shutil.rmtree(dst_path)
            else:
                dst_path.unlink()
        except OSError as e:
            return False, f"failed to remove existing dst: {e}", []
    dirs_created = []
    if parents:
        ancestors = []
        p = dst_path.parent
        while not p.exists():
            ancestors.append(p)
            p = p.parent
        ancestors.reverse()
        try:
            dst_path.parent.mkdir(parents=True, exist_ok=True)
        except PermissionError as e:
            return False, f"permission denied creating dst parent (possibly outside allowed scope): {e}", []
        except OSError as e:
            return False, f"os error creating dst parent: {e}", []
        dirs_created = ancestors
    try:
        if copy:
            shutil.copy2(str(src_path), str(dst_path))
        else:
            shutil.move(str(src_path), str(dst_path))
    except PermissionError as e:
        return False, f"permission denied (possibly outside allowed scope): {e}", []
    except OSError as e:
        return False, f"os error: {e}", []
    return True, None, dirs_created


_MOVE_ERROR_CLASSES = {
    "ERR_ARG_INVALID": "invalid_args",
    "ERR_ARG_MISSING": "invalid_args",
    "ERR_TEMPLATE_FAIL": "invalid_args",
    "ERR_AMBIGUOUS_PATH": "ambiguous_input",
    "ERR_PATH_NOT_FOUND": "not_found",
    "ERR_DST_EXISTS": "conflict",
    "ERR_REFUSE_MOVE": "unsafe_target",
    "ERR_PERMISSION_DENIED": "permission_denied",
    "ERR_OP_FAILED": "operation_failed",
}


def _move_error_code(error: str | None) -> str:
    text = str(error or "").lower()
    if text.startswith("src not found:"):
        return "ERR_PATH_NOT_FOUND"
    if text.startswith("dst already exists"):
        return "ERR_DST_EXISTS"
    if text.startswith("src and dst are the same path"):
        return "ERR_REFUSE_MOVE"
    if "permission denied" in text:
        return "ERR_PERMISSION_DENIED"
    return "ERR_OP_FAILED"


def _move_failure(error_code: str, error: str) -> dict:
    return {
        "ok": False,
        "ok_count": 0,
        "fail_count": 0,
        "results": [],
        "dirs_created": [],
        "failed": [],
        "error_class": _MOVE_ERROR_CLASSES.get(error_code, "operation_failed"),
        "error_code": error_code,
        "error": error,
    }


def _move_alias_failure(payload: dict, *, prefix: str) -> dict:
    error_code = str(payload.get("error_code") or "ERR_AMBIGUOUS_PATH")
    error = f"{prefix}{payload.get('error') or ''}"
    out = _move_failure(error_code, error)
    for key in ("candidates", "input_path", "hint"):
        if key in payload:
            out[key] = payload[key]
    return out


def _move_result(results: list, failed: list, **extra) -> dict:
    out = vector_result(results, failed, entry_key="results")
    out.update(extra)
    if failed:
        primary = failed[0]
        error_code = primary.get("error_code") or "ERR_OP_FAILED"
        out.update({
            "error_class": _MOVE_ERROR_CLASSES.get(
                error_code, "operation_failed"),
            "error_code": error_code,
            "error": primary.get("error") or _msg(
                "ERR_OP_FAILED", reason="move failed"),
        })
    return out


def move(args: dict) -> dict:
    """Sposta/rinomina entries (vettoriale). Args: entries, dst_template, ..."""
    if not isinstance(args, dict):
        return _move_failure(
            "ERR_ARG_INVALID",
            _msg("ERR_ARG_INVALID", arg="args", reason="must be an object"),
        )
    entries = args.get("entries")
    dst_template = args.get("dst_template")
    overwrite = bool(args.get("overwrite", False))
    parents = bool(args.get("parents", True))
    allow_dirs = bool(args.get("allow_dirs", False))
    allow_system = bool(args.get("allow_system", False))
    copy_mode = bool(args.get("copy", False))

    if entries is None or not isinstance(entries, list):
        return _move_failure(
            "ERR_ARG_INVALID",
            _msg("ERR_ARG_INVALID", arg="entries", reason="must be a list"),
        )
    if not dst_template or not isinstance(dst_template, str):
        return _move_failure(
            "ERR_ARG_INVALID",
            _msg("ERR_ARG_INVALID", arg="dst_template", reason="must be a string"),
        )

    # D.3: pre-check ambiguita' alias bilingue su tutti src + dst_template
    # parent. Se anche solo UN path e' ambiguo, fail globale (move never
    # implicit delete: meglio non muovere nessuno che muovere quello
    # sbagliato). Skip se dst_template ha placeholder ({...}).
    if "{" not in dst_template:
        ambig_dst = _check_mutating_path_ambiguity(
            dst_template, target_must_exist=False)
        if ambig_dst is not None:
            return _move_alias_failure(
                ambig_dst, prefix="dst_template ambiguo: ")
    for i, entry in enumerate(entries[:50]):  # cap pre-check a 50
        if not isinstance(entry, dict):
            continue
        src_arg = entry.get("path") or entry.get("src")
        if not src_arg or not isinstance(src_arg, str):
            continue
        ambig_src = _check_mutating_path_ambiguity(
            src_arg, target_must_exist=True)
        if ambig_src is not None:
            return _move_alias_failure(
                ambig_src,
                prefix=f"src[{i}] '{src_arg}' ambiguo: ",
            )

    results = []
    failed = []
    all_dirs_created = set()
    for i, entry in enumerate(entries):
        if not isinstance(entry, dict):
            failed.append({"index": i, "error_code": "ERR_ARG_INVALID",
                           "error": _msg("ERR_ARG_INVALID", arg=f"entries[{i}]", reason="must be a dict")})
            continue
        src_arg = entry.get("path") or entry.get("src")
        if not src_arg or not isinstance(src_arg, str):
            failed.append({"index": i, "error_code": "ERR_ARG_MISSING",
                           "error": _msg("ERR_ARG_MISSING", arg=f"entries[{i}].path (o 'src')")})
            continue
        src_path = Path(os.path.expanduser(src_arg)).resolve()
        kind = entry.get("kind") or ""
        if not allow_dirs and (kind == "dir" or src_path.is_dir()):
            failed.append({"index": i, "src": str(src_path),
                           "error_code": "ERR_REFUSE_MOVE",
                           "error": _msg("ERR_REFUSE_MOVE", path=str(src_path),
                                          reason="directory (passa allow_dirs=true o filtra prima con filter_entries)")})
            continue
        if not allow_system and is_system_file(src_path.name):
            failed.append({"index": i, "src": str(src_path),
                           "error_code": "ERR_REFUSE_MOVE",
                           "error": _msg("ERR_REFUSE_MOVE", path=str(src_path),
                                          reason=f"system file '{src_path.name}' (passa allow_system=true o filtra prima)")})
            continue
        try:
            fields = _entry_fields(entry, src_path)
            dst_str = dst_template.format(**fields)
        except KeyError as e:
            failed.append({"index": i, "src": str(src_path),
                           "error_code": "ERR_TEMPLATE_FAIL",
                           "error": _msg("ERR_TEMPLATE_FAIL", stage="placeholder", reason=str(e))})
            continue
        except Exception as e:
            failed.append({"index": i, "src": str(src_path),
                           "error_code": "ERR_TEMPLATE_FAIL",
                           "error": _msg("ERR_TEMPLATE_FAIL", stage="render", reason=str(e))})
            continue
        dst_path = Path(os.path.expanduser(dst_str)).resolve()
        ok, err, dirs_created = _move_one(src_path, dst_path, overwrite,
                                          parents, copy=copy_mode)
        if ok:
            results.append({"src": str(src_path), "dst": str(dst_path)})
            for d in dirs_created:
                all_dirs_created.add(str(d))
        else:
            failed.append({
                "index": i,
                "src": str(src_path),
                "dst": str(dst_path),
                "error_code": _move_error_code(err),
                "error": err,
            })

    dirs_created_list = sorted(all_dirs_created, key=lambda p: p.count("/"), reverse=True)

    return _move_result(results, failed, dirs_created=dirs_created_list)


def reverse_move(plan, results):
    """Undo multistage di move(): sposta dst→src + rimuove dir create vuote."""
    if not isinstance(plan, dict) or not isinstance(results, dict):
        return _move_failure(
            "ERR_ARG_INVALID",
            _msg("ERR_ARG_INVALID", arg="plan/results", reason="must be objects"),
        )
    pairs = (results or {}).get("results") or []
    dirs_created = (results or {}).get("dirs_created") or []
    if not isinstance(pairs, list) or not isinstance(dirs_created, list):
        return _move_failure(
            "ERR_ARG_INVALID",
            _msg(
                "ERR_ARG_INVALID", arg="results",
                reason="results and dirs_created must be lists",
            ),
        )
    out_results, failed = [], []

    for i, p in enumerate(pairs):
        if (not isinstance(p, dict) or not isinstance(p.get("src"), str)
                or not isinstance(p.get("dst"), str)):
            failed.append({
                "index": i,
                "error_code": "ERR_ARG_INVALID",
                "error": _msg(
                    "ERR_ARG_INVALID", arg=f"results[{i}]",
                    reason="src and dst strings are required",
                ),
            })
            continue
        src_now = Path(p["dst"])
        dst_back = Path(p["src"])
        ok, err, _ = _move_one(src_now, dst_back, overwrite=False, parents=True)
        if ok:
            out_results.append({"src": str(src_now), "dst": str(dst_back)})
        else:
            failed.append({
                "index": i,
                "src": str(src_now),
                "dst": str(dst_back),
                "error_code": _move_error_code(err),
                "error": err,
            })

    dirs_removed = []
    dirs_kept = []
    sorted_dirs = sorted(dirs_created, key=lambda p: p.count("/"), reverse=True)
    for d_str in sorted_dirs:
        d = Path(d_str)
        if not d.exists():
            continue
        if not d.is_dir():
            continue
        try:
            if not any(d.iterdir()):
                d.rmdir()
                dirs_removed.append(str(d))
            else:
                dirs_kept.append(str(d))
        except OSError:
            dirs_kept.append(str(d))

    return _move_result(
        out_results,
        failed,
        dirs_created=[],
        dirs_removed=dirs_removed,
        dirs_kept=dirs_kept,
    )


# --- find_dirs -------------------------------------------------------------


def find_dirs(args: dict) -> dict:
    """Walk ricorsivo dell'albero di directory con metadati aggregati."""
    base_path = args.get("base_path")
    recursive = args.get("recursive", True)
    max_depth = args.get("max_depth", 10)
    max_results = args.get("max_results", 1000)
    include_hidden = bool(args.get("include_hidden", False))

    if not base_path:
        return {"ok": False, "error_code": "ERR_ARG_MISSING",
                "error": _msg("ERR_ARG_MISSING", arg="base_path")}
    if not isinstance(max_results, int) or max_results < 1:
        return {"ok": False, "error_code": "ERR_ARG_INVALID",
                "error": _msg("ERR_ARG_INVALID", arg="max_results", reason="must be a positive integer")}
    if not isinstance(max_depth, int) or max_depth < 0:
        return {"ok": False, "error_code": "ERR_ARG_INVALID",
                "error": _msg("ERR_ARG_INVALID", arg="max_depth", reason="must be >= 0")}

    base, alias_note = _resolve_path_with_alias(base_path)
    if not base.exists():
        # Suggerisci cartelle home esistenti: il planner puo' chiedere
        # all'utente quale intendeva, evitando loop_break generico.
        return {"ok": False, "error_code": "ERR_PATH_NOT_FOUND",
                "error": _msg("ERR_PATH_NOT_FOUND", path=str(base)),
                "suggested_paths": _home_dir_suggestions(base.name)}
    if not base.is_dir():
        return {"ok": False, "error_code": "ERR_PATH_WRONG_TYPE",
                "error": _msg("ERR_PATH_WRONG_TYPE", expected="directory", actual="file", path=str(base))}

    entries: list[dict] = []
    truncated = False
    visited_dirs = 0

    def _scan_dir(d: Path) -> dict | None:
        try:
            file_count = 0
            total_bytes = 0
            size_min: int | None = None
            size_max: int | None = None
            for child in d.iterdir():
                if not include_hidden and child.name.startswith("."):
                    continue
                try:
                    if child.is_symlink():
                        continue
                    if child.is_file():
                        file_count += 1
                        s = child.stat().st_size
                        total_bytes += s
                        if size_min is None or s < size_min:
                            size_min = s
                        if size_max is None or s > size_max:
                            size_max = s
                except OSError:
                    continue
            try:
                mt = d.stat().st_mtime
            except OSError:
                mt = 0.0
            return {
                "path": str(d),
                "name": d.name,
                "file_count": file_count,
                "total_bytes": total_bytes,
                "size_min": size_min if size_min is not None else 0,
                "size_max": size_max if size_max is not None else 0,
                "mtime": float(mt),
            }
        except PermissionError:
            return None
        except OSError:
            return None

    # La base NON e' una "directory IN base" (e' il contenitore della ricerca):
    # ne misuriamo i file diretti SOLO per l'aggregato file_count_total, ma non
    # entra in `entries`. recursive=false → figli immediati (iterdir);
    # recursive=true → tutti i discendenti (rglob). Bug 31/5/2026: prima la base
    # era l'unica entry con recursive=false → "quante directory in /etc" = 1.
    base_entry = _scan_dir(base)
    base_file_count = int((base_entry or {}).get("file_count", 0) or 0)
    try:
        iterator = base.rglob("*") if recursive else base.iterdir()
        for p in iterator:
            try:
                if p.is_symlink() or not p.is_dir():
                    continue
                rel_parts = p.relative_to(base).parts
            except (ValueError, OSError):
                continue
            if recursive and len(rel_parts) > max_depth:
                continue
            if not include_hidden and any(seg.startswith(".") for seg in rel_parts):
                continue
            entry = _scan_dir(p)
            visited_dirs += 1
            if entry is None:
                continue
            entries.append(entry)
            if len(entries) >= max_results:
                truncated = True
                break
    except PermissionError as e:
        return {"ok": False,
                "error_code": "ERR_PERMISSION_DENIED",
                "error": _msg("ERR_PERMISSION_DENIED"), "detail": str(e)}
    except OSError as e:
        return {"ok": False, "error_code": "ERR_OP_FAILED",
                "error": _msg("ERR_OP_FAILED", reason=f"os error: {e}")}

    matches = [e["path"] for e in entries]
    # Aggregati anti-confusione (turn af6447da 22/5/2026): il LLM ha letto
    # `count` di find_dirs come "count file" e ha risposto "858 file" quando
    # erano 858 directories. Aggiungo nomi non ambigui:
    # - count_dirs = numero di directory trovate
    # - file_count_total = somma dei file diretti su tutte le dirs (= numero
    #   totale di file ricorsivo sotto base_path, senza dover lanciare anche
    #   find_files).
    count_dirs = len(entries)
    file_count_total = base_file_count + sum(int(e.get("file_count", 0) or 0) for e in entries)
    out = {
        "ok": True,
        "entries": entries,
        "matches": matches,
        "metadata": {
            "base_path": str(base),
            "recursive": recursive,
            "include_hidden": include_hidden,
            "count": count_dirs,            # legacy, ambiguo
            "count_dirs": count_dirs,        # esplicito
            "file_count_total": file_count_total,
            "visited_dirs": visited_dirs,
            "truncated": truncated,
            **({"alias_resolved": alias_note} if alias_note else {}),
        },
    }
    if truncated:
        out["truncated"] = True
        out["truncated_what"] = "directory"
        out["used"] = len(entries)
        out["cap_field"] = "max_results"
        out["cap_value"] = max_results
    return out


# --- create_dirs -----------------------------------------------------------


def _dir_vector_result(results, failed):
    out = {
        "ok": len(failed) == 0,
        "ok_count": len(results),
        "fail_count": len(failed),
        "results": results,
        "failed": failed,
    }
    if results and failed:
        out["partial"] = True
    if failed:
        primary = failed[0]
        code = primary.get("error_code") or "ERR_DIR_OP_FAILED"
        out.update({
            "error_class": (
                "invalid_input" if code == "ERR_ARG_INVALID"
                else "ambiguous_input" if code == "ERR_AMBIGUOUS_PATH"
                else "operation_failed"
            ),
            "error_code": code,
            "error": primary.get("error") or _msg(
                "ERR_DIR_OP_FAILED", op="directory", path=str(
                    primary.get("path") or "?"), reason="operation failed"),
        })
    return out


def _create_one(path_arg, parents, exist_ok, mode):
    target = _normalize_input_path(path_arg)
    pre_existed = target.exists()
    try:
        kwargs = {"parents": bool(parents), "exist_ok": bool(exist_ok)}
        if mode is not None:
            kwargs["mode"] = mode
        target.mkdir(**kwargs)
    except FileExistsError:
        return False, str(target), f"path already exists and is not a directory: {target}", False
    except FileNotFoundError as e:
        return False, str(target), f"missing parent (use parents=true to auto-create): {e}", False
    except PermissionError as e:
        return False, str(target), f"permission denied (possibly outside allowed scope): {e}", False
    except OSError as e:
        return False, str(target), f"os error: {e}", False
    created = (not pre_existed)
    try:
        st = target.stat()
        return True, str(target), oct(st.st_mode & 0o777), created
    except OSError:
        return True, str(target), None, created


def create_dirs(args: dict) -> dict:
    """Crea directory (vettoriale). Args: paths, parents, exist_ok, mode."""
    paths = args.get("paths")
    parents = args.get("parents", True)
    exist_ok = args.get("exist_ok", True)
    mode = args.get("mode")

    if paths is None or not isinstance(paths, list):
        return _dir_vector_result([], [{
            "index": 0, "path": paths, "error_code": "ERR_ARG_INVALID",
            "error": _msg("ERR_ARG_INVALID", arg="paths",
                          reason="must be a list"),
        }])
    if mode is not None:
        if not isinstance(mode, int) or not (0 <= mode <= 0o777):
            return _dir_vector_result([], [{
                "index": 0, "path": None, "error_code": "ERR_ARG_INVALID",
                "error": _msg("ERR_ARG_INVALID", arg="mode",
                              reason="must be an integer in 0..0o777"),
            }])

    results = []
    failed = []
    for i, p in enumerate(paths):
        if not isinstance(p, str) or not p:
            failed.append({"index": i, "path": p, "error_code": "ERR_ARG_INVALID",
                           "error": _msg("ERR_ARG_INVALID", arg="path", reason="must be a non-empty string")})
            continue
        # D.3: ambiguita' alias bilingue su parent → ERR_AMBIGUOUS_PATH.
        ambig = _check_mutating_path_ambiguity(p, target_must_exist=False)
        if ambig is not None:
            failed.append({"index": i, "path": p, **ambig})
            continue
        ok, target, info, created = _create_one(p, parents, exist_ok, mode)
        if ok:
            entry = {"path": target, "created": created}
            if info:
                entry["mode_octal"] = info
            results.append(entry)
        else:
            failed.append({"index": i, "path": target,
                           "error_code": "ERR_DIR_OP_FAILED", "error": info})

    return _dir_vector_result(results, failed)


def reverse_create_dirs(plan, results):
    """Undo: rimuove le dir create dal forward (created=true), solo se vuote."""
    entries = (results or {}).get("results") or []
    out_results, failed = [], []
    candidates = [e for e in entries if e.get("created")]
    candidates.sort(key=lambda e: len(e["path"]), reverse=True)
    for i, entry in enumerate(candidates):
        path = Path(entry["path"])
        if not path.exists():
            failed.append({"index": i, "path": str(path),
                           "error_code": "ERR_PATH_NOT_FOUND",
                           "error": _msg("ERR_PATH_NOT_FOUND", path=str(path))})
            continue
        if not path.is_dir():
            failed.append({"index": i, "path": str(path),
                           "error_code": "ERR_PATH_WRONG_TYPE",
                           "error": _msg("ERR_PATH_WRONG_TYPE", expected="directory", actual="file", path=str(path))})
            continue
        try:
            children = list(path.iterdir())
        except OSError as e:
            failed.append({"index": i, "path": str(path),
                           "error_code": "ERR_DIR_OP_FAILED",
                           "error": _msg("ERR_DIR_OP_FAILED", op="list", path=str(path), reason=str(e))})
            continue
        if children:
            failed.append({"index": i, "path": str(path),
                           "error_code": "ERR_DIR_OP_FAILED",
                           "error": _msg("ERR_DIR_OP_FAILED", op="rmdir", path=str(path),
                                          reason=f"directory non vuota ({len(children)} items): no auto-remove")})
            continue
        try:
            path.rmdir()
            out_results.append({"path": str(path), "removed": True})
        except OSError as e:
            failed.append({"index": i, "path": str(path),
                           "error_code": "ERR_DIR_OP_FAILED",
                           "error": _msg("ERR_DIR_OP_FAILED", op="rmdir", path=str(path), reason=str(e))})
    return _dir_vector_result(out_results, failed)


# --- delete_dirs -----------------------------------------------------------


def _remove_one(path_arg, if_empty_only, force):
    target = Path(os.path.expanduser(path_arg)).resolve()
    if not target.exists():
        return False, str(target), "path does not exist"
    if not target.is_dir():
        return False, str(target), "not a directory"
    try:
        children = list(target.iterdir())
    except OSError as e:
        return False, str(target), f"cannot list: {e}"
    if children and not force:
        return False, str(target), f"directory not empty ({len(children)} items); use force=true for recursive remove"
    try:
        if force and children:
            shutil.rmtree(target)
        else:
            target.rmdir()
    except PermissionError as e:
        return False, str(target), f"permission denied (possibly outside allowed scope): {e}"
    except OSError as e:
        return False, str(target), f"os error: {e}"
    return True, str(target), None


def _file_revision(value: os.stat_result) -> tuple[int, int, int, int, int]:
    """Identity and content-change signals used by destructive operations."""
    return (
        int(value.st_dev), int(value.st_ino), int(value.st_size),
        int(value.st_mtime_ns), int(value.st_ctime_ns),
    )


def _sha256_path(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(65536), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _backup_delete_blob(path: Path, blob_dir: Path) -> tuple[Path, str, os.stat_result]:
    """Create or verify an immutable content-addressed backup before unlink."""
    initial = os.lstat(path)
    if not stat.S_ISREG(initial.st_mode):
        raise OSError("source is no longer a regular file")
    blob_dir.mkdir(parents=True, exist_ok=True)
    temporary: Path | None = None
    try:
        with path.open("rb") as source:
            opened = os.fstat(source.fileno())
            if _file_revision(opened) != _file_revision(initial):
                raise OSError("source changed before backup")
            fd, temporary_name = tempfile.mkstemp(
                prefix=".metnos-blob-", dir=str(blob_dir))
            temporary = Path(temporary_name)
            digest = hashlib.sha256()
            with os.fdopen(fd, "wb") as destination:
                for chunk in iter(lambda: source.read(65536), b""):
                    digest.update(chunk)
                    destination.write(chunk)
                destination.flush()
                os.fsync(destination.fileno())
            after_read = os.fstat(source.fileno())
            if _file_revision(after_read) != _file_revision(opened):
                raise OSError("source changed while backup was being created")

        blob_sha256 = digest.hexdigest()
        blob_path = blob_dir / f"{blob_sha256}.bin"
        try:
            os.link(temporary, blob_path)
        except FileExistsError:
            if _sha256_path(blob_path) != blob_sha256:
                raise OSError("existing backup blob failed integrity check")
        return blob_path, blob_sha256, after_read
    finally:
        if temporary is not None:
            try:
                temporary.unlink()
            except FileNotFoundError:
                pass


_DELETE_ERROR_CLASSES = {
    "ERR_ARG_INVALID": "invalid_args",
    "ERR_PATH_NOT_FOUND": "not_found",
    "ERR_PATH_WRONG_TYPE": "wrong_type",
    "ERR_REFUSE_MOVE": "forbidden",
    "ERR_OP_FAILED": "io_error",
}


def _delete_failure(error_code: str, error: str, *,
                    error_class: str | None = None) -> dict:
    """Return the stable terminal envelope required by the executor standard."""
    return {
        "ok": False,
        "ok_count": 0,
        "fail_count": 0,
        "results": [],
        "failed": [],
        "error_class": error_class or _DELETE_ERROR_CLASSES.get(
            error_code, "unknown"),
        "error_code": error_code,
        "error": error,
    }


def delete_files(args: dict) -> dict:
    """Rimuove file (vettoriale, NON directory). Args: paths.

    Reversibile §2.3: ogni file rimosso viene backupped come blob in
    `<METNOS_HISTORY_DIR>/<METNOS_TURN_ID>/blob/<sha256>.bin` PRIMA
    dell'unlink. Il runtime usa `restore_blob_backup` per ripristinare.

    Safety §2.9: rifiuta paths fuori dallo scope di scrittura, rifiuta
    directory (richiede `delete_dirs` esplicito), rifiuta system files
    (override via `allow_system=true` non ancora supportato qui).

    Best-effort §7.4: ogni path e' indipendente, una failure non blocca
    le altre.
    """
    if not isinstance(args, dict):
        return _delete_failure(
            "ERR_ARG_INVALID",
            _msg("ERR_ARG_INVALID", arg="args", reason="must be an object"),
            error_class="invalid_args",
        )
    paths = args.get("paths")
    if paths is None or not isinstance(paths, list):
        return _delete_failure(
            "ERR_ARG_INVALID",
            _msg("ERR_ARG_INVALID", arg="paths", reason="must be a list"),
            error_class="invalid_args",
        )

    # Storage blob: tracciamento turn per reverse pattern §2.3.
    history_dir = os.environ.get("METNOS_HISTORY_DIR") or str(
        _C.PATH_USER_DATA / "_history")
    turn_id = os.environ.get("METNOS_TURN_ID") or "no_turn"
    blob_dir = Path(history_dir) / turn_id / "blob"

    # §2.4 dominio APERTO: tolleranza wildcard (l'LLM tende ai glob, es.
    # paths=["/tmp/dir/*"]). Espansione a SOLI file regolari — mai directory
    # implicite (§2.9). 0 match → resta il path glob e fallisce a valle con
    # ERR_PATH_NOT_FOUND onesto.
    import glob as _glob
    expanded = []
    for p in paths:
        if isinstance(p, str) and any(c in p for c in "*?["):
            hits = sorted(
                m for m in _glob.glob(os.path.expanduser(p))
                if Path(m).is_file())
            expanded.extend(hits or [p])
        else:
            expanded.append(p)
    paths = expanded

    results = []
    failed = []
    for i, p in enumerate(paths):
        if not isinstance(p, str) or not p:
            item = {"index": i, "error_code": "ERR_ARG_INVALID",
                    "error": _msg("ERR_ARG_INVALID", arg="path",
                                  reason="must be a non-empty string")}
            if isinstance(p, str):
                item["path"] = p
            failed.append(item)
            continue
        # D.3: ambiguita' alias bilingue (delete = target must exist).
        ambig = _check_mutating_path_ambiguity(p, target_must_exist=True)
        if ambig is not None:
            failed.append({"index": i, "path": p, **ambig})
            continue
        try:
            abs_path = Path(os.path.abspath(os.path.expanduser(p)))
        except OSError as e:
            failed.append({"index": i, "path": p, "error_code": "ERR_OP_FAILED",
                           "error": _msg("ERR_OP_FAILED", reason=f"path resolve: {e}")})
            continue
        if abs_path.is_symlink():
            failed.append({"index": i, "path": str(abs_path),
                           "error_code": "ERR_PATH_WRONG_TYPE",
                           "expected": "regular_file", "actual": "symlink",
                           "error": _msg("ERR_PATH_WRONG_TYPE",
                                          expected="regular file", actual="symlink",
                                          path=str(abs_path))})
            continue
        if not abs_path.exists():
            failed.append({"index": i, "path": str(abs_path),
                           "error_code": "ERR_PATH_NOT_FOUND",
                           "error": _msg("ERR_PATH_NOT_FOUND", path=str(abs_path))})
            continue
        if abs_path.is_dir():
            # `expected`/`actual` STRUTTURATI oltre al testo: il recovery
            # deterministico (§7.9, mai regex su error multi-lingua) li usa
            # per riconoscere "directory passata a un consumer di file".
            failed.append({"index": i, "path": str(abs_path),
                           "error_code": "ERR_PATH_WRONG_TYPE",
                           "expected": "file", "actual": "directory",
                           "error": _msg("ERR_PATH_WRONG_TYPE",
                                          expected="file", actual="directory", path=str(abs_path))})
            continue
        if not abs_path.is_file():
            failed.append({"index": i, "path": str(abs_path),
                           "error_code": "ERR_PATH_WRONG_TYPE",
                           "expected": "file", "actual": "special",
                           "error": _msg("ERR_PATH_WRONG_TYPE",
                                          expected="file", actual="special", path=str(abs_path))})
            continue
        # Safety net: rifiuta system file (whitelisting platform_policy).
        try:
            if is_system_file(abs_path.name):
                failed.append({"index": i, "path": str(abs_path),
                               "error_code": "ERR_REFUSE_MOVE",
                               "error": _msg("ERR_REFUSE_MOVE", path=str(abs_path),
                                              reason="system file (no allow_system override)")})
                continue
        except Exception:
            pass
        # Backup atomico e verificato prima di qualunque unlink.
        try:
            blob_path, blob_sha256, backed_up_revision = _backup_delete_blob(
                abs_path, blob_dir)
        except OSError as e:
            failed.append({"index": i, "path": str(abs_path),
                           "error_code": "ERR_OP_FAILED",
                           "error": _msg("ERR_OP_FAILED", reason=f"backup blob: {e}")})
            continue
        # Unlink dopo backup confermato (§2.9 spirit).
        try:
            current = os.lstat(abs_path)
            if _file_revision(current) != _file_revision(backed_up_revision):
                raise OSError("source changed after backup; deletion refused")
            abs_path.unlink()
        except OSError as e:
            failed.append({"index": i, "path": str(abs_path),
                           "error_code": "ERR_OP_FAILED",
                           "error": _msg("ERR_OP_FAILED", reason=f"unlink: {e}")})
            continue
        results.append({
            "path": str(abs_path), "removed": True,
            "blob_path": str(blob_path), "blob_sha256": blob_sha256,
            "restore_mode": "create",
            "mode": stat.S_IMODE(backed_up_revision.st_mode),
            "mtime_ns": int(backed_up_revision.st_mtime_ns),
            "atime_ns": int(backed_up_revision.st_atime_ns),
        })

    out = vector_result(results, failed, entry_key="results")
    if failed and not results:
        primary = failed[0]
        error_code = primary.get("error_code") or "ERR_OP_FAILED"
        out.update({
            "error_class": _DELETE_ERROR_CLASSES.get(error_code, "unknown"),
            "error_code": error_code,
            "error": primary.get("error") or _msg(
                "ERR_OP_FAILED", reason="file deletion failed"),
        })
    return out


def delete_dirs(args: dict) -> dict:
    """Rimuove directory (vettoriale). Args: paths, if_empty_only, force."""
    paths = args.get("paths")
    if_empty_only = args.get("if_empty_only", True)  # informational; force=true overrides
    force = bool(args.get("force", False))

    if paths is None or not isinstance(paths, list):
        return _dir_vector_result([], [{
            "index": 0, "path": paths, "error_code": "ERR_ARG_INVALID",
            "error": _msg("ERR_ARG_INVALID", arg="paths",
                          reason="must be a list"),
        }])

    results = []
    failed = []
    for i, p in enumerate(paths):
        if not isinstance(p, str) or not p:
            failed.append({"index": i, "path": p, "error_code": "ERR_ARG_INVALID",
                           "error": _msg("ERR_ARG_INVALID", arg="path", reason="must be a non-empty string")})
            continue
        # D.3: ambiguita' alias bilingue (delete_dirs = target must exist).
        # Rischio massimo: cancellare 33578 file NAS quando si voleva 116 in ~/images.
        ambig = _check_mutating_path_ambiguity(p, target_must_exist=True)
        if ambig is not None:
            failed.append({"index": i, "path": p, **ambig})
            continue
        ok, target, info = _remove_one(p, if_empty_only, force)
        if ok:
            results.append({"path": target, "removed": True})
        else:
            failed.append({"index": i, "path": target,
                           "error_code": "ERR_DIR_OP_FAILED", "error": info})

    return _dir_vector_result(results, failed)


# --- spreadsheet (LOCAL, default client) -----------------------------------
# §10.3 self-hosted default: lo spreadsheet canonico e' un file LOCALE (.xlsx
# via openpyxl, .csv via stdlib), allegabile a un'email. Google Sheets e' il
# backend provider-qualified opt-in (client="google_workspace"). Per il backend
# locale `spreadsheet_id` == il PATH del file (non un Drive id).

def _spreadsheet_out_dir() -> Path:
    """Directory di default per i nuovi spreadsheet locali (auto-creata)."""
    d = _C.PATH_USER_DATA / "spreadsheets"
    d.mkdir(parents=True, exist_ok=True)
    return d


def _sanitize_filename(title: str) -> str:
    """title → filename safe (no separatori di path, no caratteri riservati)."""
    name = re.sub(r"[^\w\-. ]", "_", (title or "").strip()) or "spreadsheet"
    return name[:120]


def _resolve_xlsx_path(args: dict, *, must_exist: bool):
    """Risolve il path del file spreadsheet locale da `spreadsheet_id`|`path`.

    Per il backend locale `spreadsheet_id` E' il path. `must_exist` distingue
    read/write (richiedono il file) da create (lo genera).
    """
    raw = args.get("spreadsheet_id") or args.get("path")
    if raw is not None and not isinstance(raw, str):
        return None, {"ok": False, "error_code": "ERR_ARG_NOT_STRING",
                      "error": _msg("ERR_ARG_NOT_STRING", arg="spreadsheet_id"),
                      "error_class": "invalid_args"}
    raw = (raw or "").strip()
    if not raw:
        return None, {"ok": False, "error_code": "ERR_ARG_MISSING",
                      "error": _msg("ERR_ARG_MISSING", arg="spreadsheet_id"),
                      "error_class": "invalid_args"}
    p = Path(raw).expanduser()
    if not p.is_absolute():
        p = _spreadsheet_out_dir() / p
    if must_exist and not p.is_file():
        return None, {"ok": False, "error_code": "ERR_PATH_NOT_FOUND",
                      "error": _msg("ERR_PATH_NOT_FOUND", path=str(p)),
                      "error_class": "not_found"}
    return p, None


def _normalize_rows(values) -> list:
    """values → matrice list[list] (tollera lista piatta = 1 colonna)."""
    if not isinstance(values, list):
        return []
    rows = []
    for r in values:
        rows.append(list(r) if isinstance(r, (list, tuple)) else [r])
    return rows


def _is_csv(path: Path) -> bool:
    return path.suffix.lower() == ".csv"


def _xlsx_col_name(index: int) -> str:
    out = ""
    while index:
        index, rem = divmod(index - 1, 26)
        out = chr(65 + rem) + out
    return out


def _xlsx_cell_xml(value, ref: str) -> str:
    if value is None:
        return f'<c r="{ref}"/>'
    if isinstance(value, bool):
        return f'<c r="{ref}" t="b"><v>{1 if value else 0}</v></c>'
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return f'<c r="{ref}"><v>{value}</v></c>'
    text = value if isinstance(value, str) else json.dumps(
        value, ensure_ascii=False)
    escaped = html.escape(text, quote=False)
    return (f'<c r="{ref}" t="inlineStr"><is><t xml:space="preserve">'
            f'{escaped}</t></is></c>')


def _write_xlsx_stdlib(path: Path, rows: list, sheet_name: str) -> None:
    """Scrive un XLSX minimo interoperabile usando solo zipfile/XML stdlib.

    Il file finale e' aperto in modalita' esclusiva ``xb``: una destinazione
    comparsa fra pre-check e publish non viene mai sovrascritta, anche su
    Windows dove la pubblicazione via hard-link non e' sempre disponibile.
    """
    sheet_rows: list[str] = []
    for row_no, row in enumerate(rows, 1):
        cells = "".join(_xlsx_cell_xml(value, f"{_xlsx_col_name(col_no)}{row_no}")
                        for col_no, value in enumerate(row, 1))
        sheet_rows.append(f'<row r="{row_no}">{cells}</row>')
    sheet_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        '<sheetData>' + "".join(sheet_rows) + '</sheetData></worksheet>')
    safe_sheet = html.escape(sheet_name[:31], quote=True)
    payloads = {
        "[Content_Types].xml": (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
            '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
            '<Default Extension="xml" ContentType="application/xml"/>'
            '<Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
            '<Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
            '<Override PartName="/xl/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml"/>'
            '</Types>'),
        "_rels/.rels": (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>'
            '</Relationships>'),
        "xl/workbook.xml": (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
            'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
            f'<sheets><sheet name="{safe_sheet}" sheetId="1" r:id="rId1"/></sheets>'
            '</workbook>'),
        "xl/_rels/workbook.xml.rels": (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>'
            '<Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" Target="styles.xml"/>'
            '</Relationships>'),
        "xl/styles.xml": (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
            '<numFmts count="0"/>'
            '<fonts count="1"><font><sz val="11"/><name val="Calibri"/></font></fonts>'
            '<fills count="2"><fill><patternFill patternType="none"/></fill>'
            '<fill><patternFill patternType="gray125"/></fill></fills>'
            '<borders count="1"><border><left/><right/><top/><bottom/><diagonal/></border></borders>'
            '<cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellStyleXfs>'
            '<cellXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/></cellXfs>'
            '<cellStyles count="1"><cellStyle name="Normal" xfId="0" builtinId="0"/></cellStyles>'
            '<dxfs count="0"/>'
            '<tableStyles count="0" defaultTableStyle="TableStyleMedium9" defaultPivotStyle="PivotStyleLight16"/>'
            '</styleSheet>'),
        "xl/worksheets/sheet1.xml": sheet_xml,
    }
    created = False
    try:
        with path.open("xb") as raw_file:
            created = True
            with zipfile.ZipFile(raw_file, "w", zipfile.ZIP_DEFLATED) as archive:
                for name, content in payloads.items():
                    archive.writestr(name, content.encode("utf-8"))
    except Exception:
        if created:
            try:
                path.unlink()
            except OSError:
                pass
        raise


# Sinonimi di campo bilingui per il mapping entries→colonne (§2.10): il planner
# nomina le colonne in linguaggio utente ("descrizione", "percorso") mentre le
# entries hanno chiavi tecniche ("description", "path"). Riusabile/estendibile.
_FIELD_SYNONYMS = {
    "path": ("percorso", "image_path", "filepath", "file", "src"),
    "description": ("descrizione", "desc", "caption"),
    "name": ("nome", "filename", "title", "titolo"),
    "size_bytes": ("size", "dimensione", "bytes"),
    "score": ("punteggio", "rilevanza", "relevance"),
    "keywords": ("parole_chiave", "tags", "tag"),
    "date": ("data", "datetime", "timestamp"),
    "dominio": ("domini", "domain", "domains"),
    "origine": ("origini", "origin", "origins", "source", "sources"),
}


def _entry_cell(entry: dict, col: str):
    """Valore di `entry` per la colonna `col`, risolvendo i sinonimi di campo."""
    def _render(value):
        return (", ".join(map(str, value))
                if isinstance(value, (list, tuple)) else value)

    def _usable(value) -> bool:
        return value is not None and value != "" and value != []

    if col in entry and _usable(entry[col]):
        return _render(entry[col])
    cl = col.strip().lower()
    if cl in entry and _usable(entry[cl]):
        return _render(entry[cl])
    for canon, syns in _FIELD_SYNONYMS.items():
        names = (canon,) + syns
        if cl in names:
            for n in names:
                if n in entry and _usable(entry[n]):
                    return _render(entry[n])
    return ""


def _entries_to_values(entries, columns) -> list:
    """Costruisce la matrice [header + righe] da una LISTA di entries (dict) e
    una lista di colonne (nomi di campo). Risolve la frizione list→matrice che
    il planner non sa esprimere coi placeholder (§2.10). Colonne assenti =
    chiavi non-interne della prima entry."""
    rows_in = [e for e in (entries or []) if isinstance(e, dict)]
    cols = [c for c in (columns or []) if isinstance(c, str) and c.strip()]
    if not cols and rows_in:
        cols = [k for k in rows_in[0].keys() if not str(k).startswith("_")]
    if not cols:
        return []
    out = [list(cols)]
    n = len(cols)
    for e in rows_in:
        cells = [_entry_cell(e, c) for c in cols]
        # §2.4 robustezza NL→determinismo: se la risoluzione per-chiave/sinonimo
        # lascia colonne vuote (il planner ha nominato `fields` e `columns` in
        # lingue/nomi diversi, es. date↔data, address↔indirizzo) ma l'entry ha
        # lo STESSO numero di campi → mapping POSIZIONALE (liste parallele nello
        # stesso ordine). Recupera i dati invece di celle vuote. Turn 1671283e.
        if sum(1 for c in cells if c not in ("", None)) < n:
            vals = [v for k, v in e.items() if not str(k).startswith("_")]
            if len(vals) == n:
                cells = vals
        out.append(cells)
    return out


def _rows_entries_to_values(entries, columns) -> list:
    """entries = RIGHE (list[list], es. output di read_spreadsheet §2.6) → matrice
    diretta. §2.8 (bug silent-loss, turn 3da933e5-family): `_entries_to_values`
    filtra `isinstance(e, dict)` → le righe list[list] venivano SCARTATE →
    foglio header-only, dati persi in silenzio. Qui le preserviamo. Se `columns`
    è dato e la larghezza combacia e la riga-0 NON è già quell'header → antepone
    l'header d'uscita; altrimenti copia le righe as-is (dati > cosmesi)."""
    rows = _normalize_rows(entries)
    cols = [c for c in (columns or []) if isinstance(c, str) and c.strip()]
    if cols and rows and len(rows[0]) == len(cols):
        head0 = [str(c).strip().lower() for c in rows[0]]
        if head0 != [c.strip().lower() for c in cols]:
            return [list(cols)] + rows
    return rows


def _resolve_values(args: dict):
    """Risolve la matrice di celle da `values` (diretta) o `entries`+`columns`
    (pipe §2.10/§4.1). Ritorna list[list] o None se nessuna fonte valida."""
    values = args.get("values")
    if isinstance(values, list) and values:
        return _normalize_rows(values)
    entries = args.get("entries")
    if isinstance(entries, list) and entries:
        # §2.8: entries possono essere RECORD (list[dict], da extract/list_*) o
        # RIGHE (list[list], da read_spreadsheet). Le righe NON vanno scartate:
        # matrice diretta (dati preservati), i dict via mapping colonne.
        has_dict = any(isinstance(e, dict) for e in entries)
        has_rows = any(isinstance(e, (list, tuple)) for e in entries)
        if has_rows and not has_dict:
            return _rows_entries_to_values(entries, args.get("columns"))
        return _entries_to_values(entries, args.get("columns"))
    if isinstance(values, list):  # lista vuota esplicita = foglio vuoto
        return []
    return None


def create_spreadsheet(args: dict) -> dict:
    """Crea un nuovo spreadsheet LOCALE (.xlsx default, .csv se path .csv).

    Args:
      - `title`: str (richiesto) → nome file (se `path` assente).
      - `values`: list[list] (opzionale) → righe iniziali (es. header + dati).
      - `sheet_name`: str (opzionale, default "Sheet1").
      - `path`: str (opzionale) → path esplicito; altrimenti
        `~/.local/share/metnos/spreadsheets/<title>.xlsx`.
    Output §2.6 trasformativo: `results: [{spreadsheet_id, path, title, rows}]`.
    Reverse: `delete_created_paths`.
    """
    if not isinstance(args, dict):
        return {"ok": False, "error_code": "ERR_ARG_INVALID",
                "error": _msg("ERR_ARG_INVALID", arg="args", reason="must be an object"),
                "error_class": "invalid_args", "results": [], "used": 0, "n_created": 0}
    title = (args.get("title") or "").strip()
    explicit_path = args.get("path") or args.get("spreadsheet_id")
    if not title and not explicit_path:
        # `title` NON è obbligatorio: auto-nome col timestamp. Rationale: create
        # richiedeva title (che il planner doveva INVENTARE) mentre write
        # richiede values (che la query fornisce) → la frizione spingeva il
        # planner verso write per "metti i valori in uno spreadsheet". Senza
        # title forzato, create ha la stessa bassa frizione di write.
        title = "spreadsheet_" + datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    if explicit_path:
        path = Path(str(explicit_path)).expanduser()
        if not path.is_absolute():
            path = _spreadsheet_out_dir() / path
        # §2.4: un output senza estensione nota (o path-verbatim del planner,
        # es. `C:\Windows\…` finito in `path` — turn a9ec3b06) produrrebbe un
        # file che nessun reader riapre. Suffisso .xlsx di default.
        if path.suffix.lower() not in (".xlsx", ".csv"):
            path = path.with_name(_sanitize_filename(path.name) + ".xlsx")
    else:
        path = _spreadsheet_out_dir() / f"{_sanitize_filename(title)}.xlsx"
    if path.exists():
        return {"ok": False, "error_code": "ERR_DST_EXISTS",
                "error": _msg("ERR_DST_EXISTS", path=str(path)),
                "error_class": "conflict",
                "results": [], "used": 0, "n_created": 0}
    path.parent.mkdir(parents=True, exist_ok=True)
    rows = _resolve_values(args) or []
    sheet_name = (args.get("sheet_name") or "Sheet1").strip() or "Sheet1"
    try:
        if _is_csv(path):
            import csv
            with path.open("x", newline="", encoding="utf-8") as fh:
                csv.writer(fh).writerows(rows)
        else:
            _write_xlsx_stdlib(path, rows, sheet_name)
    except FileExistsError:
        return {"ok": False, "error_code": "ERR_DST_EXISTS",
                "error": _msg("ERR_DST_EXISTS", path=str(path)),
                "error_class": "conflict",
                "results": [], "used": 0, "n_created": 0}
    except Exception as e:  # noqa: BLE001
        return {"ok": False, "error_code": "ERR_IO",
                "error": _msg("ERR_IO", detail=str(e)),
                "error_class": "io_error", "results": [], "used": 0, "n_created": 0}
    sid = str(path)
    # `created: True` + `path`: contratto richiesto da reverse_patterns.
    # _delete_created_paths (undo: il file appena creato viene rimosso, §2.8).
    result_row = {"ok": True, "created": True, "spreadsheet_id": sid, "path": sid,
                  "title": title or path.stem, "rows": len(rows), "kind": "spreadsheet"}
    return {
        "ok": True, "n_created": 1, "spreadsheet_id": sid, "path": sid,
        "title": title or path.stem, "results": [result_row], "used": 1,
        "files_source": "local",
    }


def write_spreadsheet(args: dict) -> dict:
    """Scrive/appende righe in uno spreadsheet LOCALE esistente.

    Args:
      - `spreadsheet_id`: str (richiesto) = PATH del file locale.
      - `values`: list[list] (richiesto).
      - `mode`: "overwrite" (default, riscrive il foglio) | "append".
      - `sheet_name`: str (opzionale, default primo foglio).
    Output §2.6: `results: [{spreadsheet_id, updated_cells, mode}]`.
    """
    if not isinstance(args, dict):
        return {"ok": False, "error_code": "ERR_ARG_INVALID",
                "error": _msg("ERR_ARG_INVALID", arg="args", reason="must be an object"),
                "error_class": "invalid_args", "results": [], "used": 0, "n_written": 0}
    # Valida gli ARGS prima di toccare il filesystem (no I/O su input malformato).
    # `rows` da `values` (matrice diretta) o da `entries`+`columns` (pipe §2.10):
    # il planner non sa comporre una matrice con placeholder su una LISTA, quindi
    # accettiamo le entries (from_step) + le colonne e costruiamo la matrice qui.
    rows = _resolve_values(args)
    if rows is None:
        return {"ok": False, "error_code": "ERR_ARG_INVALID",
                "error": _msg("ERR_ARG_INVALID", arg="values",
                                reason="provide `values` (matrix) or `entries`+`columns`"),
                "error_class": "invalid_args", "results": [], "used": 0, "n_written": 0}
    mode = (args.get("mode") or "overwrite").strip().lower()
    if mode not in ("overwrite", "append"):
        return {"ok": False, "error_code": "ERR_ARG_INVALID",
                "error": _msg("ERR_ARG_INVALID", arg="mode", reason="must be 'overwrite' or 'append'"),
                "error_class": "invalid_args", "results": [], "used": 0, "n_written": 0}
    # UPSERT (§2.4 robustezza NL→determinismo): `spreadsheet_id` OPZIONALE per
    # il backend locale. "metti/salva … in uno spreadsheet" presuppone
    # linguisticamente un contenitore esistente; quando NON esiste (nessun id,
    # o un path inesistente) accomodiamo la presupposizione CREANDO il file,
    # invece di fallire o chiedere l'id. Cosi' l'ambiguita' verbale create-vs-
    # write non rompe la pipeline: qualunque verbo scelga il planner, esce un
    # .xlsx locale. Se invece l'id punta a un file esistente, lo MODIFICA.
    raw = args.get("spreadsheet_id") or args.get("path")
    if raw is not None and not isinstance(raw, str):
        return {"ok": False, "error_code": "ERR_ARG_NOT_STRING",
                "error": _msg("ERR_ARG_NOT_STRING", arg="spreadsheet_id"),
                "error_class": "invalid_args", "results": [], "used": 0, "n_written": 0}
    raw = (raw or "").strip()
    if raw:
        path = Path(raw).expanduser()
        if not path.is_absolute():
            path = _spreadsheet_out_dir() / path
    else:
        path = _spreadsheet_out_dir() / (
            "spreadsheet_" + datetime.datetime.now().strftime("%Y%m%d_%H%M%S") + ".xlsx")
    path.parent.mkdir(parents=True, exist_ok=True)
    existed = path.is_file()
    # Undo §2.3: se il file ESISTE gia', backup blob dei bytes previ PRIMA di
    # modificarlo (overwrite o append), cosi' `restore_blob_backup` puo'
    # ripristinare lo stato pre-write. File NUOVO → niente blob (l'undo lo
    # rimuove via `delete_created_paths`). Stessa convenzione di `delete_files`.
    prev_blob_path = None
    if existed:
        try:
            import shutil as _shutil
            import hashlib as _hashlib
            history_dir = os.environ.get("METNOS_HISTORY_DIR") or str(
                _C.PATH_USER_DATA / "_history")
            turn_id = os.environ.get("METNOS_TURN_ID") or "no_turn"
            blob_dir = Path(history_dir) / turn_id / "blob"
            h = _hashlib.sha256()
            with path.open("rb") as _f:
                for chunk in iter(lambda: _f.read(65536), b""):
                    h.update(chunk)
            blob_sha256 = h.hexdigest()
            blob_dir.mkdir(parents=True, exist_ok=True)
            _bp = blob_dir / f"{blob_sha256}.bin"
            if not _bp.exists():
                _shutil.copy2(str(path), str(_bp))
            prev_blob_path = str(_bp)
        except OSError:
            # Backup fallito → onesti (§2.8): non dichiariamo undo per questo file.
            prev_blob_path = None
    sheet_name = (args.get("sheet_name") or "").strip()
    try:
        if _is_csv(path):
            import csv
            file_mode = "a" if (mode == "append" and existed) else "w"
            with path.open(file_mode, newline="", encoding="utf-8") as fh:
                csv.writer(fh).writerows(rows)
        else:
            import openpyxl
            if existed:
                wb = openpyxl.load_workbook(str(path))
                ws = (wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames
                      else wb.active)
                if mode == "overwrite":
                    ws.delete_rows(1, ws.max_row)
            else:
                wb = openpyxl.Workbook()
                ws = wb.active
                if sheet_name:
                    ws.title = sheet_name[:31]
            for r in rows:
                ws.append(r)
            wb.save(str(path))
    except ImportError:
        return {"ok": False, "error_code": "ERR_DEPENDENCY_MISSING",
                "error": _msg("ERR_DEPENDENCY_MISSING", what="openpyxl"),
                "error_class": "dependency_missing", "results": [], "used": 0, "n_written": 0}
    except Exception as e:  # noqa: BLE001
        return {"ok": False, "error_code": "ERR_IO",
                "error": _msg("ERR_IO", detail=str(e)),
                "error_class": "io_error", "results": [], "used": 0, "n_written": 0}
    sid = str(path)
    updated_cells = sum(len(r) for r in rows)
    created = not existed
    # Schema result §2.6 con campi-undo §2.3 letti dal catalogo:
    #   - file NUOVO → `created=true`+`path` → `delete_created_paths` lo rimuove.
    #   - file PREESISTENTE → `path`+`prev_blob_path` → `restore_blob_backup`
    #     ripristina i bytes previ (annulla overwrite/append).
    result_row = {"ok": True, "spreadsheet_id": sid, "path": sid,
                  "updated_cells": updated_cells, "mode": mode, "created": created}
    if created:
        result_row["created"] = True
    elif prev_blob_path:
        result_row["prev_blob_path"] = prev_blob_path
    out = {
        "ok": True, "n_written": 1, "updated_cells": updated_cells,
        "updated_rows": len(rows), "spreadsheet_id": sid, "path": sid,
        "mode": mode, "created": created,
        "results": [result_row], "used": 1, "files_source": "local",
    }
    # Undo onesto (§2.8): reverse_pattern multistage nel manifest
    # (`restore_blob_backup` + `delete_created_paths`); ogni stadio agisce sul
    # campo che lo riguarda (prev_blob_path vs created), saltando gli altri.
    if created:
        out["_undo"] = {"reverse_pattern": "delete_created_paths", "paths": [sid]}
    elif prev_blob_path:
        out["_undo"] = {"reverse_pattern": "restore_blob_backup",
                        "paths": [sid], "blob_path": prev_blob_path}
    return out


def append_spreadsheet(args: dict) -> dict:
    """Append-only wrapper di `write_spreadsheet` (mode='append')."""
    a = dict(args or {})
    a["mode"] = "append"
    return write_spreadsheet(a)


def read_spreadsheet(args: dict) -> dict:
    """Legge tutte le righe da uno spreadsheet LOCALE.

    Args:
      - `spreadsheet_id`: str (richiesto) = PATH del file locale.
      - `sheet_name`: str (opzionale, default primo foglio).
    Output: `{ok, values: [[...]], entries, spreadsheet_id, used}`.
    """
    if not isinstance(args, dict):
        return {"ok": False, "error_code": "ERR_ARG_INVALID",
                "error": _msg("ERR_ARG_INVALID", arg="args", reason="must be an object"),
                "error_class": "invalid_args", "entries": [], "used": 0}
    path, err = _resolve_xlsx_path(args, must_exist=True)
    if err is not None:
        return {**err, "entries": [], "used": 0}
    sheet_name = (args.get("sheet_name") or "").strip()
    try:
        if _is_csv(path):
            import csv
            with path.open("r", newline="", encoding="utf-8") as fh:
                values = [list(r) for r in csv.reader(fh)]
        else:
            import openpyxl
            wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
            ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
            values = [list(row) for row in ws.iter_rows(values_only=True)]
    except ImportError:
        return {"ok": False, "error_code": "ERR_DEPENDENCY_MISSING",
                "error": _msg("ERR_DEPENDENCY_MISSING", what="openpyxl"),
                "error_class": "dependency_missing", "entries": [], "used": 0}
    except Exception as e:  # noqa: BLE001
        return {"ok": False, "error_code": "ERR_IO",
                "error": _msg("ERR_IO", detail=str(e)),
                "error_class": "io_error", "entries": [], "used": 0}
    sid = str(path)
    return {"ok": True, "values": values, "spreadsheet_id": sid,
            "entries": values, "used": len(values),
            "available_total": len(values), "files_source": "local"}
