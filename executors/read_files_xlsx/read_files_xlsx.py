#!/usr/bin/env python3
"""read_files_xlsx — executor di Metnos v1.1.

Legge file Excel (.xlsx) e ritorna le righe come lista di dict.
Vettoriale: una sola call accetta una lista di paths.

Backend: openpyxl (read-only mode per efficienza).

Contratto:
    stdin: JSON {paths: list[str], sheet?: str|int (default first),
                 has_header?: bool = true, max_rows?: int = 10000}
    stdout: JSON {ok, ok_count, fail_count, entries, failed}
            entries[i] = {path, sheet, headers, rows: list[dict|list], row_count}
"""
import json
import os
import sys
from pathlib import Path

sys.path.insert(0, os.environ.get("METNOS_RUNTIME") or next(
    str(p / "runtime") for p in Path(__file__).resolve().parents
    if (p / "runtime" / "config.py").is_file()))
from messages import get as _msg  # noqa: E402
from executor_helpers import coerce_cap, run_stdio, vector_result  # noqa: E402

try:
    import openpyxl
except ImportError:
    openpyxl = None


def _failed(error_class, error_code, error, *, detail=None):
    out = {
        "error_class": error_class,
        "error_code": error_code,
        "error": error,
    }
    if detail:
        out["detail"] = detail
    return out


def _read_one(path_arg, sheet, has_header, max_rows):
    if openpyxl is None:
        return None, _failed(
            "dependency_missing", "openpyxl_missing",
            _msg("ERR_DEPENDENCY_MISSING", what="openpyxl"),
        )
    path = Path(os.path.expanduser(path_arg)).resolve()
    if not path.exists():
        return None, _failed(
            "not_found", "path_not_found",
            _msg("ERR_PATH_NOT_FOUND", path=str(path)),
        )
    if not path.is_file():
        return None, _failed(
            "invalid_input", "path_not_file",
            _msg("ERR_PATH_WRONG_TYPE", expected="file", actual="non-file",
                 path=str(path)),
        )
    try:
        wb = openpyxl.load_workbook(filename=str(path), read_only=True, data_only=True)
    except Exception as e:
        return None, _failed(
            "invalid_content", "xlsx_parse_failed",
            _msg("ERR_FILE_READ_FAILED", path=str(path)), detail=str(e),
        )
    try:
        if sheet is None:
            ws = wb.worksheets[0]
        elif isinstance(sheet, int):
            if sheet < 0 or sheet >= len(wb.worksheets):
                return None, _failed(
                    "invalid_input", "worksheet_index_invalid",
                    _msg("ERR_WORKSHEET_INDEX_INVALID", index=sheet,
                         maximum=len(wb.worksheets) - 1),
                )
            ws = wb.worksheets[sheet]
        else:
            if sheet not in wb.sheetnames:
                return None, _failed(
                    "not_found", "worksheet_not_found",
                    _msg("ERR_WORKSHEET_NOT_FOUND", sheet=sheet),
                    detail=", ".join(wb.sheetnames),
                )
            ws = wb[sheet]
        headers = None
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            row_list = list(row)
            if i == 0 and has_header:
                headers = [str(h).strip() if h is not None else f"col_{j}" for j, h in enumerate(row_list)]
                continue
            if has_header and headers is not None:
                rows.append({headers[j] if j < len(headers) else f"col_{j}": v for j, v in enumerate(row_list)})
            else:
                rows.append(row_list)
            if len(rows) >= max_rows:
                break
        # available_total reale: ws.max_row include header. Se cap raggiunto,
        # available_total = (max_row - 1 se has_header) per le righe dati.
        available_total = len(rows)
        truncated = False
        if len(rows) >= max_rows:
            try:
                total_data_rows = ws.max_row - (1 if has_header else 0)
                if total_data_rows > len(rows):
                    truncated = True
                    available_total = total_data_rows
            except Exception:
                pass
        return {
            "path": str(path),
            "sheet": ws.title,
            "headers": headers,
            "rows": rows,
            "row_count": len(rows),
            "all_sheets": wb.sheetnames,
            "_truncated": truncated,
            "_available_total": available_total,
        }, None
    finally:
        wb.close()


def invoke(args):
    if not isinstance(args, dict):
        return {
            "ok": False,
            "error": _msg("ERR_ARGS_NOT_OBJECT"),
            "error_class": "invalid_input",
            "error_code": "args_not_object",
        }
    paths = args.get("paths")
    sheet = args.get("sheet")
    has_header = bool(args.get("has_header", True))
    max_rows = coerce_cap(args, "max_rows", 10000, maximum=1000000)

    if not isinstance(paths, list):
        return {
            "ok": False,
            "error": _msg("ERR_ARG_NOT_LIST", arg="paths"),
            "error_class": "invalid_input",
            "error_code": "paths_not_list",
        }

    entries, failed = [], []
    aggregate_truncated = False
    aggregate_used = 0
    aggregate_available = 0
    for i, p in enumerate(paths):
        if not isinstance(p, str) or not p:
            failed.append({
                "index": i,
                "path": p,
                **_failed(
                    "invalid_input", "path_not_nonempty_string",
                    _msg("ERR_ARG_NOT_NONEMPTY_STRING", arg="path"),
                ),
            })
            continue
        entry, err = _read_one(p, sheet, has_header, max_rows)
        if err:
            failed.append({
                "index": i,
                "path": str(Path(os.path.expanduser(p)).resolve()),
                **err,
            })
            continue
        if entry.pop("_truncated", False):
            aggregate_truncated = True
        aggregate_used += entry.get("row_count", 0)
        aggregate_available += entry.pop("_available_total", entry.get("row_count", 0))
        entries.append(entry)

    out = vector_result(entries, failed)
    if aggregate_truncated:
        out["truncated"] = True
        out["truncated_what"] = _msg("MSG_OBJECT_LINES")
        out["used"] = aggregate_used
        out["available_total"] = aggregate_available
        out["cap_field"] = "max_rows"
        out["cap_value"] = max_rows
    return out


def main():
    run_stdio(invoke)


if __name__ == "__main__":
    main()
